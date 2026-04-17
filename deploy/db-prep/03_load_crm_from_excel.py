#!/usr/bin/env python3
"""
Load accounts, contacts, and opportunities from three Excel workbooks (header row = 1).

Run after 02_seed_employees_from_entra.py (opportunity owners match employees by email or name).

Requires: openpyxl, asyncpg (same as backend). DATABASE_URL without +asyncpg.
"""

from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import asyncpg
from openpyxl import load_workbook


def normalize_database_url(url: str) -> str:
    u = url.strip()
    if "+asyncpg" in u:
        u = u.replace("postgresql+asyncpg://", "postgresql://", 1)
    return u


async def pg_enum_typname(conn: asyncpg.Connection, table: str, column: str) -> str:
    row = await conn.fetchrow(
        """
        SELECT t.typname::text AS typname
        FROM pg_attribute a
        JOIN pg_class c ON a.attrelid = c.oid
        JOIN pg_namespace n ON c.relnamespace = n.oid
        JOIN pg_type t ON a.atttypid = t.oid
        WHERE n.nspname = 'public'
          AND c.relname = $1
          AND a.attname = $2
          AND a.attnum > 0
          AND NOT a.attisdropped
        """,
        table,
        column,
    )
    if not row:
        raise RuntimeError(f"Column not found: public.{table}.{column}")
    name = row["typname"]
    if not re.fullmatch(r"[a-z0-9_]+", name):
        raise RuntimeError(f"Unexpected type name: {name!r}")
    return name


def norm_cell(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v).strip() or None


def header_map(ws) -> dict[str, int]:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    m: dict[str, int] = {}
    for i, raw in enumerate(row1):
        key = norm_cell(raw)
        if key:
            m[key.lower()] = i
    return m


def col_idx(m: dict[str, int], *candidates: str) -> int | None:
    for c in candidates:
        j = m.get(c.lower())
        if j is not None:
            return j
    return None


def parse_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = norm_cell(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def parse_decimal(v: Any) -> Decimal | None:
    if v is None:
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = norm_cell(v)
    if not s:
        return None
    s = s.replace(",", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


ACCOUNT_TYPE_ALIASES: dict[str, str] = {
    "customer": "customer",
    "client": "customer",
    "vendor": "vendor",
    "supplier": "vendor",
    "partner": "partner",
    "network": "network",
    "prospect": "customer",
}

OPP_STATUS_ALIASES: dict[str, str] = {
    "discovery": "discovery",
    "qualification": "discovery",
    "qualified": "qualified",
    "qualify": "qualified",
    "proposal": "proposal",
    "proposed": "proposal",
    "negotiation": "negotiation",
    "negotiating": "negotiation",
    "won": "won",
    "closed won": "won",
    "lost": "lost",
    "closed lost": "lost",
    "cancelled": "cancelled",
    "canceled": "cancelled",
    "pipeline": "discovery",
    "active": "negotiation",
    "in progress": "negotiation",
    "progress": "negotiation",
}

ACCOUNTABILITY_ALIASES: dict[str, str] = {
    "full ownership": "full_ownership",
    "full_ownership": "full_ownership",
    "mgmt accountable": "mgmt_accountable",
    "management accountable": "mgmt_accountable",
    "mgmt_accountable": "mgmt_accountable",
    "mgmt advisory": "mgmt_advisory",
    "management advisory": "mgmt_advisory",
    "mgmt_advisory": "mgmt_advisory",
    "staff aug limited": "staff_aug_limited",
    "staff aug": "staff_aug_limited",
    "staff_aug_limited": "staff_aug_limited",
}

STRATEGIC_ALIASES: dict[str, str] = {
    "critical": "critical",
    "high": "high",
    "medium": "medium",
    "med": "medium",
    "low": "low",
}

# Default US mapping matches common seed (North America); override with OPPORTUNITY_SEED_DELIVERY_CENTER.
_DEFAULT_US_DC = os.environ.get("OPPORTUNITY_SEED_DELIVERY_CENTER", "North America").strip() or "North America"

# Map free-text country / region (Excel) to delivery_centers.name in the database
COUNTRY_TO_DC_NAME: dict[str, str] = {
    "united states": _DEFAULT_US_DC,
    "usa": _DEFAULT_US_DC,
    "us": _DEFAULT_US_DC,
    "u.s.": _DEFAULT_US_DC,
    "u.s.a.": _DEFAULT_US_DC,
    "canada": "Canada",
    "united kingdom": "United Kingdom",
    "uk": "United Kingdom",
    "great britain": "United Kingdom",
    "india": "India",
    "germany": "Germany",
    "france": "France",
    "australia": "Australia",
    "singapore": "Singapore",
    "mexico": "Mexico",
    "brazil": "Brazil",
    "japan": "Japan",
    "ireland": "Ireland",
    "netherlands": "Netherlands",
    "spain": "Spain",
    "italy": "Italy",
    "sweden": "Sweden",
    "switzerland": "Switzerland",
    "philippines": "Philippines",
    "poland": "Poland",
}

# Map country-like strings to ISO currency for accounts (fallback USD)
COUNTRY_TO_CURRENCY: dict[str, str] = {
    "united states": "USD",
    "usa": "USD",
    "us": "USD",
    "canada": "CAD",
    "united kingdom": "GBP",
    "uk": "GBP",
    "india": "INR",
    "germany": "EUR",
    "france": "EUR",
    "ireland": "EUR",
    "netherlands": "EUR",
    "spain": "EUR",
    "italy": "EUR",
    "australia": "AUD",
    "japan": "JPY",
    "singapore": "SGD",
    "mexico": "MXN",
    "brazil": "BRL",
    "switzerland": "CHF",
    "sweden": "SEK",
    "poland": "PLN",
    "philippines": "PHP",
}


def country_key(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def resolve_dc_name(country_cell: str | None) -> str:
    if not country_cell:
        return _DEFAULT_US_DC
    k = country_key(country_cell)
    if k in COUNTRY_TO_DC_NAME:
        return COUNTRY_TO_DC_NAME[k]
    # Allow exact match to an existing DC name (case-insensitive) — caller validates against DB
    return country_cell.strip()


def country_to_currency(country_cell: str | None) -> str:
    if not country_cell:
        return "USD"
    k = country_key(country_cell)
    return COUNTRY_TO_CURRENCY.get(k, "USD")


def map_account_type(raw: str | None) -> str:
    if not raw:
        return "customer"
    k = raw.strip().lower()
    return ACCOUNT_TYPE_ALIASES.get(k, k if k in ("customer", "vendor", "partner", "network") else "customer")


VALID_OPP_STATUS = frozenset(
    {"discovery", "qualified", "proposal", "negotiation", "won", "lost", "cancelled"}
)


def map_opp_status(raw: str | None) -> str:
    if not raw:
        return "discovery"
    k = raw.strip().lower()
    v = OPP_STATUS_ALIASES.get(k, k)
    return v if v in VALID_OPP_STATUS else "discovery"


ALLOW_ACCOUNTABILITY = frozenset(
    {"full_ownership", "mgmt_accountable", "mgmt_advisory", "staff_aug_limited"}
)
ALLOW_STRATEGIC = frozenset({"critical", "high", "medium", "low"})


def map_accountability(raw: str | None) -> str | None:
    if not raw:
        return None
    k = raw.strip().lower()
    v = ACCOUNTABILITY_ALIASES.get(k)
    return v if v in ALLOW_ACCOUNTABILITY else None


def map_strategic(raw: str | None) -> str | None:
    if not raw:
        return None
    k = raw.strip().lower()
    v = STRATEGIC_ALIASES.get(k)
    return v if v in ALLOW_STRATEGIC else None


# --- Opportunity deal / forecast (mirror app.services.opportunity_service.OpportunityService) ---

CLOSING_OPPORTUNITY_STATUS = frozenset({"won", "lost", "cancelled"})

# calculate_probability_from_status — DB stores lowercase status strings
PROBABILITY_BY_STATUS: dict[str, float] = {
    "discovery": 10.0,
    "qualified": 25.0,
    "proposal": 50.0,
    "negotiation": 80.0,
    "won": 100.0,
}

# app.utils.currency_converter._get_default_rates (used when currency_rates is empty)
_DEFAULT_CURRENCY_RATES_TO_USD: dict[str, float] = {
    "USD": 1.0,
    "PHP": 50.0,
    "VND": 24000.0,
    "THB": 35.0,
    "EUR": 0.85,
    "GBP": 0.75,
    "AUD": 1.35,
    "SGD": 1.35,
    "JPY": 110.0,
    "CNY": 6.5,
}


async def load_currency_rates_to_usd(conn: asyncpg.Connection) -> dict[str, float]:
    """Uppercase currency_code -> rate_to_usd (units of currency per 1 USD), same semantics as backend."""
    rows = await conn.fetch("SELECT upper(trim(currency_code::text)) AS c, rate_to_usd FROM currency_rates")
    if not rows:
        return dict(_DEFAULT_CURRENCY_RATES_TO_USD)
    return {str(r["c"]): float(r["rate_to_usd"]) for r in rows}


def probability_from_opp_status(status: str) -> float:
    """Match OpportunityService.calculate_probability_from_status (lost/cancelled -> 0)."""
    return float(PROBABILITY_BY_STATUS.get(status, 0.0))


def calculate_forecast_value(probability: float | None, deal_value: Decimal | None) -> Decimal | None:
    """Match OpportunityService.calculate_forecast_value: deal * (probability / 100)."""
    if probability is None or deal_value is None:
        return None
    return Decimal(str(float(deal_value) * (probability / 100.0)))


def calculate_deal_value_usd(
    deal_value: Decimal | None,
    currency: str,
    rates: dict[str, float],
) -> Decimal | None:
    """Match OpportunityService.calculate_deal_value_usd using currency_rates (amount / rate)."""
    if deal_value is None:
        return None
    cu = (currency or "USD").strip().upper() or "USD"
    if cu == "USD":
        return deal_value
    rate = rates.get(cu, 1.0)
    if rate == 0:
        return None
    return Decimal(str(float(deal_value) / rate))


def resolve_close_date_for_import(status: str, close_from_sheet: date | None) -> date | None:
    """Closing rows get a close_date (sheet value or today); open rows keep optional sheet close."""
    if status in CLOSING_OPPORTUNITY_STATUS:
        return close_from_sheet if close_from_sheet is not None else date.today()
    return close_from_sheet


def calculate_deal_length_days(creation_date: date | None, close_date: date | None) -> int | None:
    """Match OpportunityService.calculate_deal_length."""
    if creation_date is None:
        return None
    end_date = close_date if close_date is not None else date.today()
    if end_date < creation_date:
        return 0
    return (end_date - creation_date).days


async def load_reference(conn: asyncpg.Connection) -> tuple[Any, dict[str, Any], dict[str, Any]]:
    bt = await conn.fetchrow("SELECT id FROM billing_terms WHERE code = $1", "NET30")
    if bt is None:
        raise RuntimeError("billing_terms row with code NET30 not found.")
    dc_rows = await conn.fetch("SELECT id, lower(name) AS lk, name FROM delivery_centers")
    dc_by_lower = {r["lk"]: r["id"] for r in dc_rows}

    emp_rows = await conn.fetch(
        "SELECT id, lower(email) AS lem, lower(trim(first_name || ' ' || last_name)) AS lname "
        "FROM employees"
    )
    by_email = {r["lem"]: r["id"] for r in emp_rows if r["lem"]}
    by_name = {r["lname"]: r["id"] for r in emp_rows if r["lname"]}
    dc_exact = {r["name"]: r["id"] for r in dc_rows}
    return bt["id"], dc_by_lower, {"by_email": by_email, "by_name": by_name, "dc_exact": dc_exact}


def resolve_dc_id(country_cell: str | None, dc_by_lower: dict[str, Any], dc_exact: dict[str, Any]) -> Any:
    name = resolve_dc_name(country_cell)
    lid = name.strip().lower()
    if lid in dc_by_lower:
        return dc_by_lower[lid]
    for k, vid in dc_exact.items():
        if k.lower() == lid:
            return vid
    raise ValueError(f"No delivery_centers row for invoice country / center {name!r}")


def resolve_owner_id(owner_raw: str | None, emp: dict[str, Any]) -> Any | None:
    if not owner_raw:
        return None
    s = owner_raw.strip()
    le = s.lower()
    if "@" in le:
        return emp["by_email"].get(le)
    ln = re.sub(r"\s+", " ", le).strip()
    return emp["by_name"].get(ln)


async def run(
    dsn: str,
    clients_path: Path,
    contacts_path: Path,
    deals_path: Path,
) -> None:
    conn = await asyncpg.connect(dsn)
    try:
        typ_account = await pg_enum_typname(conn, "accounts", "type")
        typ_status = await pg_enum_typname(conn, "opportunities", "status")
        typ_acc = await pg_enum_typname(conn, "opportunities", "accountability")
        typ_strat = await pg_enum_typname(conn, "opportunities", "strategic_importance")

        billing_id, dc_by_lower, emp_ctx = await load_reference(conn)
        currency_rates = await load_currency_rates_to_usd(conn)
        dc_exact = emp_ctx["dc_exact"]
        emp = {"by_email": emp_ctx["by_email"], "by_name": emp_ctx["by_name"]}

        # --- Accounts ---
        wb = load_workbook(clients_path, read_only=True, data_only=True)
        ws = wb.active
        hm = header_map(ws)
        ix_name = col_idx(hm, "name")
        ix_type = col_idx(hm, "type")
        ix_ind = col_idx(hm, "industry")
        ix_country = col_idx(hm, "country")
        if ix_name is None:
            raise RuntimeError(f"Clients sheet missing Name column. Found headers: {list(hm)}")

        account_sql = f"""
INSERT INTO accounts (
    id, company_name, type, industry, street_address, city, region, country,
    billing_term_id, default_currency, created_at
) VALUES (
    $1, $2, $3::{typ_account}, $4, NULL, NULL, NULL, $5,
    $6, $7, now()
)
ON CONFLICT (company_name) DO UPDATE SET
    type = EXCLUDED.type,
    industry = EXCLUDED.industry,
    country = EXCLUDED.country,
    billing_term_id = EXCLUDED.billing_term_id,
    default_currency = EXCLUDED.default_currency
RETURNING id
"""

        company_to_id: dict[str, Any] = {}
        n_acc = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = norm_cell(row[ix_name] if ix_name < len(row) else None)
            if not name:
                continue
            typ = map_account_type(norm_cell(row[ix_type] if ix_type is not None and ix_type < len(row) else None))
            ind = norm_cell(row[ix_ind] if ix_ind is not None and ix_ind < len(row) else None)
            country = norm_cell(row[ix_country] if ix_country is not None and ix_country < len(row) else None) or "Unknown"
            cur = country_to_currency(country)
            aid = uuid.uuid4()
            rid = await conn.fetchval(
                account_sql,
                aid,
                name,
                typ,
                ind,
                country,
                billing_id,
                cur,
            )
            company_to_id[name.strip().lower()] = rid
            n_acc += 1
        wb.close()
        print(f"Accounts processed: {n_acc}", file=sys.stderr)

        # --- Contacts ---
        wb = load_workbook(contacts_path, read_only=True, data_only=True)
        ws = wb.active
        hm = header_map(ws)
        ix_acct = col_idx(hm, "account")
        ix_fn = col_idx(hm, "first name")
        ix_ln = col_idx(hm, "last name")
        ix_em = col_idx(hm, "email")
        ix_ph = col_idx(hm, "phone")
        ix_job = col_idx(hm, "job title")
        if ix_acct is None or ix_fn is None or ix_ln is None:
            raise RuntimeError(f"Contacts sheet missing required columns. Found: {list(hm)}")

        contact_sql = """
INSERT INTO contacts (
    id, account_id, first_name, last_name, email, phone, job_title, is_primary, is_billing
) VALUES ($1, $2, $3, $4, $5, $6, $7, 'false', 'false')
"""
        n_ct = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            def g(i: int | None) -> Any:
                if i is None or i >= len(row):
                    return None
                return row[i]

            acct_name = norm_cell(g(ix_acct))
            if not acct_name:
                continue
            aid = company_to_id.get(acct_name.strip().lower())
            if aid is None:
                print(f"SKIP contact: unknown account {acct_name!r}", file=sys.stderr)
                continue
            fn = norm_cell(g(ix_fn)) or "Unknown"
            ln = norm_cell(g(ix_ln)) or "Unknown"
            em = norm_cell(g(ix_em))
            ph = norm_cell(g(ix_ph))
            job = norm_cell(g(ix_job))
            await conn.execute(
                contact_sql,
                uuid.uuid4(),
                aid,
                fn,
                ln,
                em,
                ph,
                job,
            )
            n_ct += 1
        wb.close()
        print(f"Contacts inserted: {n_ct}", file=sys.stderr)

        # --- Opportunities ---
        wb = load_workbook(deals_path, read_only=True, data_only=True)
        ws = wb.active
        hm = header_map(ws)
        ix_name = col_idx(hm, "name")
        ix_acct = col_idx(hm, "account")
        ix_state = col_idx(hm, "state")
        ix_country = col_idx(hm, "country")
        ix_cur = col_idx(hm, "currency")
        ix_sd = col_idx(hm, "start date")
        ix_ed = col_idx(hm, "end date")
        ix_owner = col_idx(hm, "owner")
        ix_dc = col_idx(hm, "deal creation date")
        ix_acc = col_idx(hm, "accountability")
        ix_str = col_idx(hm, "strategic importance")
        ix_close = col_idx(hm, "close date")
        ix_val = col_idx(hm, "deal value")
        if ix_name is None or ix_acct is None:
            raise RuntimeError(f"Deal sheet missing Name or Account. Found: {list(hm)}")

        opp_sql_both = f"""
INSERT INTO opportunities (
    id, name, parent_opportunity_id, account_id, start_date, end_date, status,
    billing_term_id, description, utilization, margin, default_currency, delivery_center_id,
    opportunity_owner_id, invoice_customer, billable_expenses, attributes,
    accountability, strategic_importance, deal_creation_date, deal_value, deal_value_usd,
    close_date, deal_length, probability, forecast_value, forecast_value_usd
) VALUES (
    $1, $2, NULL, $3, $4, $5, $6::{typ_status},
    $7, NULL, NULL, NULL, $8, $9,
    $10, TRUE, TRUE, '{{}}'::json,
    $11::{typ_acc}, $12::{typ_strat}, $13, $14, $15,
    $16, $17, $18, $19, $20
)
"""

        opp_sql_acc_only = f"""
INSERT INTO opportunities (
    id, name, parent_opportunity_id, account_id, start_date, end_date, status,
    billing_term_id, description, utilization, margin, default_currency, delivery_center_id,
    opportunity_owner_id, invoice_customer, billable_expenses, attributes,
    accountability, strategic_importance, deal_creation_date, deal_value, deal_value_usd,
    close_date, deal_length, probability, forecast_value, forecast_value_usd
) VALUES (
    $1, $2, NULL, $3, $4, $5, $6::{typ_status},
    $7, NULL, NULL, NULL, $8, $9,
    $10, TRUE, TRUE, '{{}}'::json,
    $11::{typ_acc}, NULL, $12, $13, $14,
    $15, $16, $17, $18, $19
)
"""

        opp_sql_strat_only = f"""
INSERT INTO opportunities (
    id, name, parent_opportunity_id, account_id, start_date, end_date, status,
    billing_term_id, description, utilization, margin, default_currency, delivery_center_id,
    opportunity_owner_id, invoice_customer, billable_expenses, attributes,
    accountability, strategic_importance, deal_creation_date, deal_value, deal_value_usd,
    close_date, deal_length, probability, forecast_value, forecast_value_usd
) VALUES (
    $1, $2, NULL, $3, $4, $5, $6::{typ_status},
    $7, NULL, NULL, NULL, $8, $9,
    $10, TRUE, TRUE, '{{}}'::json,
    NULL, $11::{typ_strat}, $12, $13, $14,
    $15, $16, $17, $18, $19
)
"""

        opp_sql_neither = f"""
INSERT INTO opportunities (
    id, name, parent_opportunity_id, account_id, start_date, end_date, status,
    billing_term_id, description, utilization, margin, default_currency, delivery_center_id,
    opportunity_owner_id, invoice_customer, billable_expenses, attributes,
    accountability, strategic_importance, deal_creation_date, deal_value, deal_value_usd,
    close_date, deal_length, probability, forecast_value, forecast_value_usd
) VALUES (
    $1, $2, NULL, $3, $4, $5, $6::{typ_status},
    $7, NULL, NULL, NULL, $8, $9,
    $10, TRUE, TRUE, '{{}}'::json,
    NULL, NULL, $11, $12, $13,
    $14, $15, $16, $17, $18
)
"""

        n_op = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            def g(i: int | None) -> Any:
                if i is None or i >= len(row):
                    return None
                return row[i]

            oname = norm_cell(g(ix_name))
            acct_name = norm_cell(g(ix_acct))
            if not oname or not acct_name:
                continue
            aid = company_to_id.get(acct_name.strip().lower())
            if aid is None:
                print(f"SKIP opportunity: unknown account {acct_name!r} ({oname})", file=sys.stderr)
                continue
            st = map_opp_status(norm_cell(g(ix_state)))
            country = norm_cell(g(ix_country))
            cur = norm_cell(g(ix_cur)) or country_to_currency(country)
            sd = parse_date(g(ix_sd))
            ed = parse_date(g(ix_ed))
            if not sd or not ed:
                print(f"SKIP opportunity {oname!r}: missing start or end date", file=sys.stderr)
                continue
            if ed < sd:
                ed, sd = sd, ed
            try:
                dc_id = resolve_dc_id(country, dc_by_lower, dc_exact)
            except ValueError as e:
                print(f"SKIP opportunity {oname!r}: {e}", file=sys.stderr)
                continue
            owner_id = resolve_owner_id(norm_cell(g(ix_owner)), emp)
            dcd = parse_date(g(ix_dc))
            acc_v = map_accountability(norm_cell(g(ix_acc)))
            str_v = map_strategic(norm_cell(g(ix_str)))
            close_d = parse_date(g(ix_close))
            dv = parse_decimal(g(ix_val))

            cur_code = (cur or "USD").strip()
            prob = probability_from_opp_status(st)
            dv_usd = calculate_deal_value_usd(dv, cur_code, currency_rates)
            fv = calculate_forecast_value(prob, dv)
            fv_usd = calculate_forecast_value(prob, dv_usd)
            eff_close = resolve_close_date_for_import(st, close_d)
            dlen = calculate_deal_length_days(dcd, eff_close)

            oid = uuid.uuid4()
            if acc_v and str_v:
                await conn.execute(
                    opp_sql_both,
                    oid,
                    oname,
                    aid,
                    sd,
                    ed,
                    st,
                    billing_id,
                    cur,
                    dc_id,
                    owner_id,
                    acc_v,
                    str_v,
                    dcd,
                    dv,
                    dv_usd,
                    eff_close,
                    dlen,
                    prob,
                    fv,
                    fv_usd,
                )
            elif acc_v:
                await conn.execute(
                    opp_sql_acc_only,
                    oid,
                    oname,
                    aid,
                    sd,
                    ed,
                    st,
                    billing_id,
                    cur,
                    dc_id,
                    owner_id,
                    acc_v,
                    dcd,
                    dv,
                    dv_usd,
                    eff_close,
                    dlen,
                    prob,
                    fv,
                    fv_usd,
                )
            elif str_v:
                await conn.execute(
                    opp_sql_strat_only,
                    oid,
                    oname,
                    aid,
                    sd,
                    ed,
                    st,
                    billing_id,
                    cur,
                    dc_id,
                    owner_id,
                    str_v,
                    dcd,
                    dv,
                    dv_usd,
                    eff_close,
                    dlen,
                    prob,
                    fv,
                    fv_usd,
                )
            else:
                await conn.execute(
                    opp_sql_neither,
                    oid,
                    oname,
                    aid,
                    sd,
                    ed,
                    st,
                    billing_id,
                    cur,
                    dc_id,
                    owner_id,
                    dcd,
                    dv,
                    dv_usd,
                    eff_close,
                    dlen,
                    prob,
                    fv,
                    fv_usd,
                )
            n_op += 1
        wb.close()
        print(f"Opportunities inserted: {n_op}", file=sys.stderr)
    finally:
        await conn.close()


async def async_main() -> int:
    root = Path(__file__).resolve().parents[2]
    p = argparse.ArgumentParser(description="Load CRM data from Excel into Postgres.")
    p.add_argument(
        "--clients",
        type=Path,
        default=root / "uploads" / "Clients_1776399772.xlsx",
    )
    p.add_argument(
        "--contacts",
        type=Path,
        default=root / "uploads" / "Contacts_1776400628.xlsx",
    )
    p.add_argument(
        "--deals",
        type=Path,
        default=root / "uploads" / "Deal_Tracker_1776401286.xlsx",
    )
    args = p.parse_args()

    for label, path in ("clients", args.clients), ("contacts", args.contacts), ("deals", args.deals):
        if not path.is_file():
            print(f"Missing {label} file: {path}", file=sys.stderr)
            return 2

    dsn = os.environ.get("DATABASE_URL", "").strip()
    if not dsn:
        print("Missing DATABASE_URL", file=sys.stderr)
        return 2
    dsn = normalize_database_url(dsn)

    await run(
        dsn,
        args.clients.resolve(),
        args.contacts.resolve(),
        args.deals.resolve(),
    )
    return 0


def main() -> None:
    raise SystemExit(asyncio.run(async_main()))


if __name__ == "__main__":
    main()
