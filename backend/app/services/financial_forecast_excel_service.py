"""Excel export/import for Financial Forecast (template version must match for import)."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.financial_forecast.definition import FINANCIAL_FORECAST_DEFINITION_VERSION
from app.services.financial_forecast_service import FinancialForecastService


EXCEL_TEMPLATE_VERSION = str(FINANCIAL_FORECAST_DEFINITION_VERSION)


class FinancialForecastExcelService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def export_workbook(
        self,
        *,
        delivery_center_id: UUID,
        start_week: date,
        end_week: date,
        metric: str,
    ) -> BytesIO:
        data = await FinancialForecastService(self.session).get_forecast(
            delivery_center_id=delivery_center_id,
            start_week=start_week,
            end_week=end_week,
            metric=metric,
        )
        wb = Workbook()
        meta = wb.active
        meta.title = "Meta"
        meta["A1"] = "template_version"
        meta["B1"] = EXCEL_TEMPLATE_VERSION
        meta["A2"] = "delivery_center_id"
        meta["B2"] = str(delivery_center_id)
        meta["A3"] = "metric"
        meta["B3"] = metric
        meta["A4"] = "start_week"
        meta["B4"] = start_week.isoformat()
        meta["A5"] = "end_week"
        meta["B5"] = end_week.isoformat()

        ws = wb.create_sheet("Forecast")
        months = data["months"]
        rows = data["rows"]
        cells = data["cells"]

        headers = ["row_key", "label"] + [m["month_key"] for m in months]
        ws.append(headers)
        for r in rows:
            rk = r["row_key"]
            lab = r["label"]
            row_vals: list[Any] = [rk, lab]
            for m in months:
                mk = m["month_key"]
                row_vals.append(cells.get(rk, {}).get(mk, {}).get("value"))
            ws.append(row_vals)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    async def import_workbook(
        self,
        *,
        delivery_center_id: UUID,
        employee_id: UUID,
        file_stream: BytesIO,
    ) -> dict[str, Any]:
        wb = load_workbook(file_stream)
        if "Meta" not in wb.sheetnames or "Forecast" not in wb.sheetnames:
            raise ValueError("Invalid workbook: missing Meta or Forecast sheet. Export a new file from the app.")
        meta = wb["Meta"]
        tv = str(meta["B1"].value or "")
        if tv != EXCEL_TEMPLATE_VERSION:
            raise ValueError(f"template_version mismatch (file={tv}, required={EXCEL_TEMPLATE_VERSION}). Export a new file.")
        file_dc = str(meta["B2"].value or "")
        if file_dc != str(delivery_center_id):
            raise ValueError("delivery_center_id mismatch. Export a new file.")

        ws = wb["Forecast"]
        header = [c.value for c in ws[1]]
        if not header or header[0] != "row_key" or header[1] != "label":
            raise ValueError("Invalid Forecast header row.")

        month_keys = [str(x) for x in header[2:] if x]
        current = await FinancialForecastService(self.session).get_forecast(
            delivery_center_id=delivery_center_id,
            start_week=date.fromisoformat(str(meta["B4"].value)),
            end_week=date.fromisoformat(str(meta["B5"].value)),
            metric=str(meta["B3"].value or "forecast"),
        )
        cur_cells = current["cells"]

        expense_patches: list[dict[str, Any]] = []
        override_patches: list[dict[str, Any]] = []

        static_defs = {r["row_key"]: r for r in current["rows"]}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            rk = str(row[0])
            if rk not in static_defs and not rk.startswith("expense:"):
                raise ValueError(f"Unknown row_key in Excel: {rk}. Export a new file.")
            meta_row = static_defs.get(rk) or {}
            for i, mk in enumerate(month_keys):
                col_idx = 2 + i
                if col_idx >= len(row):
                    break
                raw_val = row[col_idx]
                if raw_val is None or raw_val == "":
                    continue
                try:
                    new_v = float(raw_val)
                except (TypeError, ValueError):
                    continue
                old_cell = cur_cells.get(rk, {}).get(mk, {})
                old_v = float(old_cell.get("value") or 0)
                if abs(new_v - old_v) < 1e-9:
                    continue
                if rk.startswith("expense:"):
                    lid = UUID(rk.split(":", 1)[1])
                    y, m = map(int, mk.split("-", 1))
                    expense_patches.append(
                        {"line_id": str(lid), "month_start_date": date(y, m, 1).isoformat(), "amount": new_v}
                    )
                elif meta_row.get("auto_row"):
                    y, m = map(int, mk.split("-", 1))
                    override_patches.append(
                        {"row_key": rk, "month_start_date": date(y, m, 1).isoformat(), "amount": new_v}
                    )
                elif meta_row.get("manual_expense") and not rk.startswith("expense:"):
                    line = next(
                        (
                            r
                            for r in current["rows"]
                            if r["row_key"] == rk and r.get("expense_line_id")
                        ),
                        None,
                    )
                    if not line or not line.get("expense_line_id"):
                        continue
                    y, m = map(int, mk.split("-", 1))
                    expense_patches.append(
                        {
                            "line_id": line["expense_line_id"],
                            "month_start_date": date(y, m, 1).isoformat(),
                            "amount": new_v,
                        }
                    )

        await FinancialForecastService(self.session).apply_bulk_patch(
            delivery_center_id=delivery_center_id,
            employee_id=employee_id,
            body={"expense_cells": expense_patches, "overrides": override_patches, "correlation_id": "excel-import"},
        )
        return {"updated_expense_cells": len(expense_patches), "updated_overrides": len(override_patches)}
