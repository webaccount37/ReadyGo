"""
Excel import service for estimates.
"""

import logging
from typing import List, Dict, Optional, Tuple
from uuid import UUID
from datetime import date, datetime
from decimal import Decimal
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from openpyxl import load_workbook

from app.db.repositories.estimate_repository import EstimateRepository
from app.db.repositories.estimate_line_item_repository import EstimateLineItemRepository
from app.db.repositories.estimate_weekly_hours_repository import EstimateWeeklyHoursRepository
from app.db.repositories.delivery_center_repository import DeliveryCenterRepository
from app.db.repositories.employee_repository import EmployeeRepository
from app.db.repositories.role_repository import RoleRepository
from app.db.repositories.role_rate_repository import RoleRateRepository
from app.db.repositories.opportunity_repository import OpportunityRepository
from app.models.estimate import EstimateLineItem, EstimateWeeklyHours
from app.models.delivery_center import DeliveryCenter
from app.models.employee import Employee
from app.models.role import Role
from app.models.role_rate import RoleRate
from app.utils.currency_converter import convert_currency

logger = logging.getLogger(__name__)


class ExcelImportService:
    """Service for importing estimates from Excel."""
    
    def __init__(self, session: AsyncSession):
        self.session = session
        self.estimate_repo = EstimateRepository(session)
        self.line_item_repo = EstimateLineItemRepository(session)
        self.weekly_hours_repo = EstimateWeeklyHoursRepository(session)
        self.delivery_center_repo = DeliveryCenterRepository(session)
        self.employee_repo = EmployeeRepository(session)
        self.role_repo = RoleRepository(session)
        self.role_rate_repo = RoleRateRepository(session)
        self.opportunity_repo = OpportunityRepository(session)
    
    async def import_estimate_from_excel(self, estimate_id: UUID, file_path: str) -> Dict:
        """Import estimate data from Excel file."""
        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        
        # Read metadata
        if "Metadata" not in wb.sheetnames:
            raise ValueError("Invalid template: Metadata sheet not found")
        
        metadata_ws = wb["Metadata"]
        metadata = self._read_metadata(metadata_ws)
        
        # Validate template
        if str(metadata["estimate_id"]) != str(estimate_id):
            raise ValueError(f"Template estimate_id ({metadata['estimate_id']}) does not match requested estimate_id ({estimate_id})")
        
        # Get estimate and opportunity
        estimate = await self.estimate_repo.get(estimate_id)
        if not estimate:
            raise ValueError("Estimate not found")
        
        opportunity = await self.opportunity_repo.get(estimate.opportunity_id)
        if not opportunity:
            raise ValueError("Opportunity not found")
        
        if str(metadata["opportunity_delivery_center_id"]) != str(opportunity.delivery_center_id):
            raise ValueError("Opportunity Invoice Center mismatch")
        
        # Read data sheet
        if "Estimate Data" not in wb.sheetnames:
            raise ValueError("Invalid template: Estimate Data sheet not found")
        
        data_ws = wb["Estimate Data"]
        
        # Validate week columns match
        expected_weeks = metadata["week_start_dates"]
        actual_weeks = self._extract_week_columns(data_ws, len(expected_weeks))
        
        if len(actual_weeks) != len(expected_weeks):
            raise ValueError(f"Week column count mismatch: expected {len(expected_weeks)}, found {len(actual_weeks)}. "
                           f"Please ensure you're importing the correct template for this estimate.")
        
        for idx, (expected, actual) in enumerate(zip(expected_weeks, actual_weeks)):
            if expected != actual:
                raise ValueError(f"Week column {idx + 1} mismatch: expected {expected.isoformat()}, found {actual.isoformat()}. "
                               f"Please ensure you're importing the correct template for this estimate.")
        
        # Parse data rows
        line_items_data = self._parse_data_rows(
            data_ws, 
            expected_weeks,
            metadata,
            opportunity.delivery_center_id,
            opportunity.default_currency or "USD"
        )
        
        # Upsert line items
        results = await self._upsert_line_items(estimate_id, line_items_data, expected_weeks)
        
        return results
    
    def _read_metadata(self, ws) -> Dict:
        """Read metadata from metadata sheet."""
        metadata = {}
        
        # Scan more rows to find all sections (employees can be beyond row 20)
        # Scan up to row 200 to be safe
        for row in range(1, 200):
            key = ws[f"A{row}"].value
            value = ws[f"B{row}"].value
            
            if not key:  # Skip empty rows
                continue
            
            if key == "estimate_id":
                metadata["estimate_id"] = UUID(value)
            elif key == "opportunity_id":
                metadata["opportunity_id"] = UUID(value)
            elif key == "opportunity_delivery_center_id":
                metadata["opportunity_delivery_center_id"] = UUID(value)
            elif key == "week_start_dates":
                metadata["week_start_dates"] = [date.fromisoformat(d) for d in value.split(",") if d]
            elif key == "phases":
                metadata["phases"] = []
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    phase_str = ws[f"B{row_idx}"].value
                    parts = phase_str.split("|")
                    if len(parts) >= 4:
                        metadata["phases"].append({
                            "name": parts[0],
                            "start_date": date.fromisoformat(parts[1]),
                            "end_date": date.fromisoformat(parts[2]),
                            "color": parts[3],
                        })
                    row_idx += 1
            elif key == "delivery_centers":
                metadata["delivery_centers"] = {}
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    dc_str = ws[f"B{row_idx}"].value
                    parts = dc_str.split("|")
                    if len(parts) >= 2:
                        metadata["delivery_centers"][parts[1]] = UUID(parts[0])
                    row_idx += 1
            elif key == "roles":
                metadata["roles"] = {}
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    role_str = ws[f"B{row_idx}"].value
                    parts = role_str.split("|")
                    if len(parts) >= 2:
                        metadata["roles"][parts[1]] = UUID(parts[0])
                    row_idx += 1
            elif key == "employees":
                metadata["employees"] = {}
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    emp_str = ws[f"B{row_idx}"].value
                    parts = emp_str.split("|")
                    if len(parts) >= 2:
                        emp_name = parts[1].strip()  # Normalize whitespace
                        emp_id = UUID(parts[0])
                        metadata["employees"][emp_name] = emp_id
                        logger.debug(f"Loaded employee from metadata: '{emp_name}' -> {emp_id}")
                    row_idx += 1
                logger.info(f"Loaded {len(metadata['employees'])} employees from metadata: {list(metadata['employees'].keys())}")
        
        # Verify we found employees
        if "employees" not in metadata:
            logger.warning("No 'employees' section found in metadata sheet!")
            metadata["employees"] = {}
        elif len(metadata["employees"]) == 0:
            logger.warning("Employees section found but empty in metadata sheet!")
        
        return metadata
    
    def _extract_week_columns(self, ws, expected_count: int) -> List[date]:
        """Extract week start dates from week header row (row 3 - column headers)."""
        weeks = []
        col = 12  # Start after fixed columns (Payable Center, Role, Employee, Cost, Rate, Cost Daily, Rate Daily, Start Date, End Date, Billable, Billable %)
        
        for idx in range(expected_count):
            cell = ws[f"{self._get_column_letter(col + idx)}3"]  # Row 3 is now the header row with week dates
            if cell.value:
                # Parse date from cell value
                if isinstance(cell.value, datetime):
                    week_date = cell.value.date()
                elif isinstance(cell.value, date):
                    week_date = cell.value
                else:
                    # Try parsing string (format: MM/DD/YYYY)
                    try:
                        week_date = datetime.strptime(str(cell.value), "%m/%d/%Y").date()
                    except:
                        # Try other formats
                        try:
                            week_date = datetime.strptime(str(cell.value), "%Y-%m-%d").date()
                        except:
                            week_date = None
                
                if week_date:
                    weeks.append(week_date)
        
        return weeks
    
    def _get_column_letter(self, col: int) -> str:
        """Get Excel column letter from 1-based column number."""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col)
    
    def _parse_data_rows(self, ws, weeks: List[date], metadata: Dict, 
                        opportunity_delivery_center_id: UUID, currency: str) -> List[Dict]:
        """Parse data rows from worksheet."""
        line_items = []
        start_row = 4  # Data starts at row 4 (row 3 is headers)
        
        # Find end of data (look for empty row or totals row)
        row = start_row
        while True:
            # Check if row is empty or contains "TOTALS"
            payable_center = ws[f"A{row}"].value
            if payable_center == "TOTALS" or (payable_center is None and row > start_row):
                break
            
            if payable_center is None:
                row += 1
                continue
            
            # Parse row data
            try:
                line_item = self._parse_line_item_row(ws, row, weeks, metadata, opportunity_delivery_center_id, currency)
                if line_item:
                    line_items.append(line_item)
                else:
                    logger.debug(f"Skipping row {row}: parsed as None (empty row)")
            except Exception as e:
                error_msg = f"Error parsing row {row}: {e}"
                logger.error(error_msg, exc_info=True)
                # Continue with next row - don't fail entire import for one bad row
                # The error will be logged but we continue processing other rows
            
            row += 1
        
        logger.info(f"Parsed {len(line_items)} line items from Excel (processed rows {start_row} to {row-1})")
        return line_items
    
    def _parse_line_item_row(self, ws, row: int, weeks: List[date], metadata: Dict,
                            opportunity_delivery_center_id: UUID, currency: str) -> Optional[Dict]:
        """Parse a single line item row."""
        # Payable Center (Column A)
        payable_center_name = ws[f"A{row}"].value
        if not payable_center_name or str(payable_center_name).strip() == "":
            return None  # Skip empty rows
        
        # Skip rows that are clearly not data (e.g., "TOTALS" label)
        if str(payable_center_name).strip().upper() == "TOTALS":
            return None
        
        delivery_centers = metadata.get("delivery_centers", {})
        delivery_center_id = delivery_centers.get(payable_center_name)
        if not delivery_center_id:
            raise ValueError(f"Row {row}: Invalid Payable Center '{payable_center_name}'")
        
        # Role (Column B)
        role_name = ws[f"B{row}"].value
        if not role_name:
            raise ValueError(f"Row {row}: Role is required")
        
        roles = metadata.get("roles", {})
        role_id = roles.get(role_name)
        if not role_id:
            raise ValueError(f"Row {row}: Invalid Role '{role_name}'")
        
        # Verify role has relationship with opportunity delivery center
        # This will be checked during upsert
        
        # Employee (Column C) - optional
        employee_name_raw = ws[f"C{row}"].value
        employee_id = None
        if employee_name_raw:
            employee_name = str(employee_name_raw).strip()  # Normalize whitespace
            employees = metadata.get("employees", {})
            
            logger.debug(f"Row {row}: Looking for employee '{employee_name}' (repr: {repr(employee_name)}) in {len(employees)} available employees")
            
            # Strategy 1: Try exact match first (with normalized whitespace)
            employee_id = employees.get(employee_name)
            if employee_id:
                logger.debug(f"Row {row}: Found exact match for employee '{employee_name}' -> {employee_id}")
            else:
                # Strategy 2: Try case-insensitive match with normalized whitespace
                for emp_name, emp_id in employees.items():
                    if emp_name.strip().lower() == employee_name.lower():
                        employee_id = emp_id
                        logger.info(f"Row {row}: Matched employee '{employee_name}' to '{emp_name}' (case-insensitive) -> {employee_id}")
                        break
                
                if not employee_id:
                    # Strategy 3: Try matching with normalized whitespace (collapse multiple spaces)
                    normalized_excel_name = " ".join(employee_name.split())
                    for emp_name, emp_id in employees.items():
                        normalized_meta_name = " ".join(emp_name.split())
                        if normalized_meta_name.lower() == normalized_excel_name.lower():
                            employee_id = emp_id
                            logger.info(f"Row {row}: Matched employee '{employee_name}' to '{emp_name}' (normalized whitespace) -> {employee_id}")
                            break
                
                if not employee_id:
                    # Log available employees for debugging
                    available_employees = list(employees.keys())
                    logger.error(f"Row {row}: Employee '{employee_name}' NOT FOUND in metadata!")
                    logger.error(f"Row {row}: Employee name from Excel (repr): {repr(employee_name)}")
                    logger.error(f"Row {row}: Employee name length: {len(employee_name)}")
                    logger.error(f"Row {row}: Available employees ({len(available_employees)}): {available_employees}")
                    # Log each available employee with repr for comparison
                    for emp_name in available_employees[:10]:
                        logger.error(f"Row {row}:   Available: '{emp_name}' (repr: {repr(emp_name)}, len: {len(emp_name)})")
                    # Check for similar names (substring match)
                    similar_found = False
                    for emp_name in available_employees:
                        if employee_name.lower() in emp_name.lower() or emp_name.lower() in employee_name.lower():
                            logger.error(f"Row {row}: Similar name found: '{emp_name}' (might be a match)")
                            similar_found = True
                    # CRITICAL: Don't reset employee to None - this is a data loss bug
                    # Instead, we should fail the import or at least warn strongly
                    raise ValueError(f"Row {row}: Employee '{employee_name}' not found in metadata. This would cause data loss. Available: {available_employees[:5]}")
        
        # Cost (Column D) - optional, will use defaults if None
        cost_value = ws[f"D{row}"].value
        cost = None
        if cost_value is not None:
            try:
                cost = Decimal(str(cost_value))
            except (ValueError, TypeError):
                # Invalid value, treat as None to use defaults
                cost = None
        
        # Rate (Column E) - optional, will use defaults if None
        rate_value = ws[f"E{row}"].value
        rate = None
        if rate_value is not None:
            try:
                rate = Decimal(str(rate_value))
            except (ValueError, TypeError):
                # Invalid value, treat as None to use defaults
                rate = None
        
        # Cost Daily (Column F) - optional, ignored (calculated from Cost)
        # Rate Daily (Column G) - optional, ignored (calculated from Rate)
        
        # Start Date (Column H)
        start_date_value = ws[f"H{row}"].value
        if not start_date_value:
            raise ValueError(f"Row {row}: Start Date is required")
        if isinstance(start_date_value, datetime):
            start_date = start_date_value.date()
        elif isinstance(start_date_value, date):
            start_date = start_date_value
        else:
            try:
                start_date = datetime.strptime(str(start_date_value), "%Y-%m-%d").date()
            except:
                raise ValueError(f"Row {row}: Invalid Start Date format")
        
        # End Date (Column I)
        end_date_value = ws[f"I{row}"].value
        if not end_date_value:
            raise ValueError(f"Row {row}: End Date is required")
        if isinstance(end_date_value, datetime):
            end_date = end_date_value.date()
        elif isinstance(end_date_value, date):
            end_date = end_date_value
        else:
            try:
                end_date = datetime.strptime(str(end_date_value), "%Y-%m-%d").date()
            except:
                raise ValueError(f"Row {row}: Invalid End Date format")
        
        if start_date > end_date:
            raise ValueError(f"Row {row}: Start Date must be <= End Date")
        
        # Billable (Column J)
        billable_value = ws[f"J{row}"].value
        billable = True
        if billable_value:
            billable_str = str(billable_value).strip().lower()
            billable = billable_str in ["yes", "true", "1", "y"]
        
        # Billable % (Column K) - Excel stores percentages as 0-1 (e.g., 0.15 for 15%)
        billable_pct_value = ws[f"K{row}"].value
        billable_pct = Decimal("0")
        if billable_pct_value is not None:
            pct_decimal = Decimal(str(billable_pct_value))
            # Excel percentage format stores 0-1, but user might enter 0-100
            # Check if value is > 1, if so assume it's 0-100 format, otherwise 0-1 format
            if pct_decimal > 1:
                # Value is in 0-100 format, convert to 0-1 for storage
                billable_pct = pct_decimal
                if billable_pct > 100:
                    raise ValueError(f"Row {row}: Billable % must be between 0 and 100")
            else:
                # Value is in 0-1 format (Excel percentage), convert to 0-100 for storage
                billable_pct = pct_decimal * 100
                if billable_pct < 0 or billable_pct > 100:
                    raise ValueError(f"Row {row}: Billable % must be between 0 and 100")
        
        # Weekly hours (Columns L onwards)
        weekly_hours = {}
        week_col_start = 12
        for idx, week in enumerate(weeks):
            col = week_col_start + idx
            col_letter = self._get_column_letter(col)
            hours_value = ws[f"{col_letter}{row}"].value
            
            if hours_value is not None:
                hours = Decimal(str(hours_value))
                if hours < 0:
                    raise ValueError(f"Row {row}, Week {week.isoformat()}: Hours must be >= 0")
                weekly_hours[week] = hours
        
        return {
            "delivery_center_id": delivery_center_id,
            "role_id": role_id,
            "employee_id": employee_id,
            "cost": cost,
            "rate": rate,
            "currency": currency,
            "start_date": start_date,
            "end_date": end_date,
            "billable": billable,
            "billable_expense_percentage": billable_pct,
            "weekly_hours": weekly_hours,
        }
    
    async def _upsert_line_items(self, estimate_id: UUID, line_items_data: List[Dict], 
                                 weeks: List[date]) -> Dict:
        """Upsert line items from parsed data.
        
        Matching strategy: Simple positional matching
        - Excel row N (data starts at row 4) matches estimate line item with row_order = N-4 (0-indexed)
        - This is a direct overwrite: row 4 in Excel updates row_order 0, row 5 updates row_order 1, etc.
        - If there are more Excel rows than existing line items, create new ones
        - If there are fewer Excel rows than existing line items, delete the extra ones
        """
        # Get existing line items sorted by row_order
        existing_line_items = await self.line_item_repo.list_by_estimate(estimate_id)
        # Sort by row_order to ensure consistent ordering
        existing_line_items = sorted(existing_line_items, key=lambda li: li.row_order if li.row_order is not None else 999999)
        
        # Create a map by row_order for easy lookup
        existing_by_row_order = {li.row_order: li for li in existing_line_items if li.row_order is not None}
        
        logger.info(f"Import: Found {len(existing_line_items)} existing line items")
        logger.info(f"Import: Mapped {len(existing_by_row_order)} line items by row_order: {sorted(existing_by_row_order.keys())}")
        for row_order, li in sorted(existing_by_row_order.items()):
            logger.debug(f"  row_order {row_order} -> line_item {li.id}")
        
        created_count = 0
        updated_count = 0
        deleted_count = 0
        errors = []
        
        # Track which line items were matched during processing
        matched_line_item_ids = set()
        
        # Get estimate for currency
        estimate = await self.estimate_repo.get(estimate_id)
        opportunity = await self.opportunity_repo.get(estimate.opportunity_id)
        opportunity_delivery_center_id = opportunity.delivery_center_id
        
        # Get max row_order for new items
        max_order = await self.line_item_repo.get_max_row_order(estimate_id)
        next_order = max_order + 1
        
        # Process each Excel row - match by position (row_order = Excel row index)
        for idx, item_data in enumerate(line_items_data):
            excel_row_number = idx + 4  # Excel rows start at 4 (row 1-3 are headers)
            row_order = idx  # 0-indexed row_order matches Excel row position
            try:
                # Verify role has relationship with opportunity delivery center
                # Check if ANY role rate exists for this role + delivery center (currency doesn't matter for this check)
                # Use first() instead of scalar_one_or_none() to handle multiple currencies gracefully
                role_rate_result = await self.session.execute(
                    select(RoleRate).where(
                        RoleRate.role_id == item_data["role_id"],
                        RoleRate.delivery_center_id == opportunity_delivery_center_id,
                    ).limit(1)
                )
                role_rate_for_lookup = role_rate_result.scalars().first()
                
                if not role_rate_for_lookup:
                    raise ValueError(f"Row {idx + 4}: Role does not have relationship with Opportunity Invoice Center")
                
                # IMPORTANT: Payable Center is reference-only and NOT used for rate determinations
                # All rate lookups must use Opportunity Invoice Center
                
                # Look up RoleRate for Opportunity Invoice Center (for rate calculations)
                opportunity_role_rate_result = await self.session.execute(
                    select(RoleRate).where(
                        RoleRate.role_id == item_data["role_id"],
                        RoleRate.delivery_center_id == opportunity_delivery_center_id,
                        RoleRate.default_currency == item_data["currency"],
                    ).limit(1)
                )
                opportunity_role_rate = opportunity_role_rate_result.scalars().first()
                
                if not opportunity_role_rate:
                    # Estimates should NEVER create RoleRate records
                    # If RoleRate doesn't exist for Opportunity Invoice Center, raise an error
                    raise ValueError(
                        f"Row {idx + 4}: RoleRate not found for Role '{item_data['role_id']}', "
                        f"Opportunity Invoice Center '{opportunity_delivery_center_id}', Currency '{item_data['currency']}'. "
                        f"Please create the RoleRate association first before using it in Estimates."
                    )
                
                # Calculate default rates if cost/rate are None
                final_cost = item_data["cost"]
                final_rate = item_data["rate"]
                
                if final_cost is None or final_rate is None:
                    # Get default rates using Opportunity Invoice Center RoleRate
                    # This is the correct source for rate lookups per user requirements
                    default_rate, default_cost = await self._get_default_rates_from_role_rate(
                        opportunity_role_rate.id,
                        item_data["role_id"],
                        opportunity_delivery_center_id,  # Use Opportunity Invoice Center for rate lookups
                        item_data["employee_id"],
                        item_data["currency"],
                    )
                    
                    if final_rate is None:
                        final_rate = default_rate
                    if final_cost is None:
                        final_cost = default_cost
                
                # Update Opportunity Invoice Center RoleRate if cost/rate changed (only if Excel provided explicit values)
                # Note: This updates the role_rate defaults, which may affect other line items
                # This is intentional - if user changes rates in Excel, they want to update the defaults
                # Only update if Excel provided explicit values (not defaults)
                if item_data["cost"] is not None and item_data["rate"] is not None:
                    cost_changed = abs(float(opportunity_role_rate.internal_cost_rate) - float(item_data["cost"])) > 0.01
                    rate_changed = abs(float(opportunity_role_rate.external_rate) - float(item_data["rate"])) > 0.01
                    
                    if cost_changed or rate_changed:
                        logger.info(f"Updating Opportunity Invoice Center role_rate {opportunity_role_rate.id} defaults: cost={item_data['cost']}, rate={item_data['rate']}")
                        opportunity_role_rate.internal_cost_rate = float(item_data["cost"])
                        opportunity_role_rate.external_rate = float(item_data["rate"])
                        await self.session.flush()
                # If Excel values were None, we used defaults but don't update role_rate defaults
                
                # Payable Center is stored separately for reference/export purposes
                # We need to look it up to validate it exists, but we don't use it for rate calculations
                payable_center_role_rate_result = await self.session.execute(
                    select(RoleRate).where(
                        RoleRate.role_id == item_data["role_id"],
                        RoleRate.delivery_center_id == item_data["delivery_center_id"],  # Payable Center
                        RoleRate.default_currency == item_data["currency"],
                    ).limit(1)
                )
                payable_center_role_rate = payable_center_role_rate_result.scalars().first()
                
                # Note: We don't require Payable Center RoleRate to exist - it's just for reference
                # But if it doesn't exist, we can't export it properly, so warn about it
                if not payable_center_role_rate:
                    logger.warning(
                        f"Row {idx + 4}: Payable Center RoleRate not found for Role '{item_data['role_id']}', "
                        f"Delivery Center '{item_data['delivery_center_id']}', Currency '{item_data['currency']}'. "
                        f"Payable Center will not be exported correctly."
                    )
                
                # Simple positional matching: Excel row N matches line item with row_order = N-4 (0-indexed)
                # Excel row 4 (first data row) -> row_order 0
                # Excel row 5 -> row_order 1, etc.
                line_item = existing_by_row_order.get(row_order)
                
                logger.info(f"Excel row {excel_row_number}: Looking for row_order {row_order}, found: {line_item.id if line_item else 'None'}")
                
                if line_item:
                    logger.info(f"Excel row {excel_row_number} (row_order {row_order}) matches existing line item {line_item.id} - updating")
                    # Update existing line item - direct overwrite with Excel values
                    await self.line_item_repo.update(
                        line_item.id,
                        role_rates_id=opportunity_role_rate.id,  # Use Opportunity Invoice Center RoleRate
                        payable_center_id=item_data["delivery_center_id"],  # Payable Center from Excel (reference only)
                        employee_id=item_data["employee_id"],
                        rate=final_rate,  # Use Excel value or default
                        cost=final_cost,  # Use Excel value or default
                        start_date=item_data["start_date"],
                        end_date=item_data["end_date"],
                        billable=item_data["billable"],
                        billable_expense_percentage=item_data["billable_expense_percentage"],
                        row_order=row_order,  # Preserve row_order position
                    )
                    updated_count += 1
                    matched_line_item_ids.add(line_item.id)  # Track as matched
                    logger.info(f"Updated line item {line_item.id} from Excel row {excel_row_number} (row_order={row_order}, rate={final_rate}, cost={final_cost})")
                else:
                    # Create new line item at this position
                    line_item = await self.line_item_repo.create(
                        estimate_id=estimate_id,
                        role_rates_id=opportunity_role_rate.id,  # Use Opportunity Invoice Center RoleRate
                        payable_center_id=item_data["delivery_center_id"],  # Payable Center from Excel (reference only)
                        employee_id=item_data["employee_id"],
                        rate=final_rate,  # Use Excel value or default
                        cost=final_cost,  # Use Excel value or default
                        currency=item_data["currency"],
                        start_date=item_data["start_date"],
                        end_date=item_data["end_date"],
                        row_order=row_order,  # Use position-based row_order
                        billable=item_data["billable"],
                        billable_expense_percentage=item_data["billable_expense_percentage"],
                    )
                    created_count += 1
                    matched_line_item_ids.add(line_item.id)  # Track as matched (newly created)
                    logger.info(f"Created new line item {line_item.id} from Excel row {excel_row_number} (row_order={row_order}, rate={final_rate}, cost={final_cost})")
                
                await self.session.flush()
                
                # Update weekly hours - always replace with Excel values
                # Delete existing weekly hours for this line item
                await self.weekly_hours_repo.delete_by_line_item(line_item.id)
                
                # Create new weekly hours from Excel data
                # Only create entries for weeks with hours > 0
                for week, hours in item_data["weekly_hours"].items():
                    if hours > 0:
                        await self.weekly_hours_repo.create(
                            estimate_line_item_id=line_item.id,
                            week_start_date=week,
                            hours=hours,
                        )
                
                logger.debug(f"Updated weekly hours for line item {line_item.id}: {len(item_data['weekly_hours'])} weeks")
                
            except Exception as e:
                error_msg = f"Row {idx + 4}: {str(e)}"
                errors.append(error_msg)
                logger.error(error_msg, exc_info=True)
        
        # Delete line items that weren't matched (removed from Excel)
        # Only delete if we successfully processed at least one row (to avoid accidental mass deletion)
        if len(line_items_data) > 0:
            all_existing_ids = {li.id for li in existing_line_items}
            unmatched_ids = all_existing_ids - matched_line_item_ids
            
            if unmatched_ids:
                logger.info(f"Found {len(unmatched_ids)} line items not present in Excel - will delete")
                for line_item_id in unmatched_ids:
                    try:
                        # Delete weekly hours first (cascade should handle this, but be explicit)
                        await self.weekly_hours_repo.delete_by_line_item(line_item_id)
                        # Delete the line item
                        await self.line_item_repo.delete(line_item_id)
                        deleted_count += 1
                        logger.info(f"Deleted line item {line_item_id} (not found in Excel)")
                    except Exception as e:
                        error_msg = f"Failed to delete line item {line_item_id}: {str(e)}"
                        errors.append(error_msg)
                        logger.error(error_msg, exc_info=True)
        
        await self.session.commit()
        
        return {
            "created": created_count,
            "updated": updated_count,
            "deleted": deleted_count,
            "errors": errors,
        }
    
    async def _get_default_rates_from_role_rate(
        self,
        role_rates_id: Optional[UUID],
        role_id: UUID,
        delivery_center_id: UUID,  # This is the opportunity delivery center ID (Invoice Center)
        employee_id: Optional[UUID],
        target_currency: str,
    ) -> Tuple[Decimal, Decimal]:
        """Get default rate and cost from a role_rate.
        
        Priority:
        1. Employee rates (if employee_id provided) - only cost, NOT rate
        2. RoleRate rates
        
        Args:
            role_rates_id: ID of the role_rate to use for rate lookup (may be None)
            role_id: Role ID (used if role_rates_id is None)
            delivery_center_id: Opportunity delivery center ID (Invoice Center) - used if role_rates_id is None, and for comparison with employee delivery center
            employee_id: Optional employee ID - if provided, only cost is taken from employee
            target_currency: Target currency for conversion
        
        Returns:
            Tuple of (rate, cost)
        """
        # Get RoleRate first to get the rate
        role_rate = None
        if role_rates_id:
            role_rate = await self.role_rate_repo.get(role_rates_id)
        else:
            # Try to find role_rate by role_id, delivery_center_id, currency
            result = await self.session.execute(
                select(RoleRate).where(
                    RoleRate.role_id == role_id,
                    RoleRate.delivery_center_id == delivery_center_id,
                    RoleRate.default_currency == target_currency,
                ).limit(1)
            )
            role_rate = result.scalars().first()
        
        if not role_rate:
            # No RoleRate found - return zeros
            # Note: Role model doesn't have rate attributes, so we can't fall back to role defaults
            # If a RoleRate is required, the caller should check for this and raise an error
            return Decimal("0"), Decimal("0")
        
        # Rate always comes from RoleRate (not employee)
        rate = Decimal(str(role_rate.external_rate))
        cost = Decimal(str(role_rate.internal_cost_rate))
        rate_currency = role_rate.default_currency
        
        # If employee is provided, use employee cost (but NOT rate)
        if employee_id:
            employee = await self.employee_repo.get(employee_id)
            if employee:
                # Compare Opportunity Invoice Center with Employee Delivery Center
                centers_match = delivery_center_id == employee.delivery_center_id if (delivery_center_id and employee.delivery_center_id) else False
                
                if centers_match:
                    # Centers match: use internal_cost_rate with NO currency conversion
                    employee_cost = Decimal(str(employee.internal_cost_rate))
                    cost = employee_cost
                else:
                    # Centers don't match: use internal_bill_rate with currency conversion
                    employee_cost = Decimal(str(employee.internal_bill_rate))
                    employee_currency = employee.default_currency or "USD"
                    
                    # Convert employee cost to target currency if needed
                    if target_currency and employee_currency.upper() != target_currency.upper():
                        employee_cost_decimal = await convert_currency(
                            float(employee_cost),
                            employee_currency,
                            target_currency,
                            self.session
                        )
                        cost = Decimal(str(employee_cost_decimal))
                    else:
                        cost = employee_cost
        
        # Convert rate to target currency if needed (only if we didn't already convert cost from employee)
        if target_currency and rate_currency.upper() != target_currency.upper():
            rate = Decimal(str(await convert_currency(float(rate), rate_currency, target_currency, self.session)))
            # Only convert cost if it came from role_rate (not employee)
            if not employee_id:
                cost = Decimal(str(await convert_currency(float(cost), rate_currency, target_currency, self.session)))
        
        return rate, cost

