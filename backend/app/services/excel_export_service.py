"""
Excel export service for estimates.
"""

import logging
import io
from typing import List, Dict, Optional
from uuid import UUID
from datetime import date, timedelta
from decimal import Decimal
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter

from app.db.repositories.estimate_repository import EstimateRepository
from app.db.repositories.delivery_center_repository import DeliveryCenterRepository
from app.db.repositories.employee_repository import EmployeeRepository
from app.db.repositories.role_repository import RoleRepository
from app.db.repositories.engagement_repository import EngagementRepository
from app.models.estimate import Estimate, EstimateLineItem, EstimateWeeklyHours, EstimatePhase
from app.models.delivery_center import DeliveryCenter
from app.models.employee import Employee
from app.models.role import Role
from app.models.role_rate import RoleRate

logger = logging.getLogger(__name__)


class ExcelExportService:
    """Service for exporting estimates to Excel."""
    
    def __init__(self, session: AsyncSession):
        self.session = session
        self.estimate_repo = EstimateRepository(session)
        self.delivery_center_repo = DeliveryCenterRepository(session)
        self.employee_repo = EmployeeRepository(session)
        self.role_repo = RoleRepository(session)
        self.engagement_repo = EngagementRepository(session)
    
    async def export_estimate_to_excel(self, estimate_id: UUID) -> io.BytesIO:
        """Export an estimate to Excel format with validation rules."""
        # Get estimate with all relationships
        estimate = await self.estimate_repo.get_with_line_items(estimate_id)
        if not estimate:
            raise ValueError("Estimate not found")
        
        # Get engagement for delivery center and currency
        engagement = await self.engagement_repo.get(estimate.engagement_id)
        if not engagement:
            raise ValueError("Engagement not found")
        
        engagement_delivery_center_id = engagement.delivery_center_id
        if not engagement_delivery_center_id:
            raise ValueError("Engagement Invoice Center (delivery_center_id) is required")
        
        # Get all delivery centers, employees, and roles for dropdowns
        all_delivery_centers = await self.delivery_center_repo.list_all()
        all_employees = await self.employee_repo.list(skip=0, limit=10000)
        
        # Get roles filtered by engagement delivery center
        roles_result = await self.session.execute(
            select(Role)
            .join(RoleRate, Role.id == RoleRate.role_id)
            .where(RoleRate.delivery_center_id == engagement_delivery_center_id)
            .distinct()
        )
        filtered_roles = list(roles_result.scalars().all())
        
        # Generate weeks based on estimate line items
        # If no line items, use engagement dates or default range
        if estimate.line_items and len(estimate.line_items) > 0:
            weeks = self._generate_weeks_from_line_items(estimate.line_items)
        else:
            # Use engagement dates or default range
            if engagement.start_date and engagement.end_date:
                weeks = self._generate_weeks_from_dates(engagement.start_date, engagement.end_date)
            else:
                # Default to 1 year from today
                from datetime import datetime
                today = datetime.now().date()
                start = today
                end = date(today.year + 1, today.month, today.day)
                weeks = self._generate_weeks_from_dates(start, end)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Estimate Data"
        
        # Create metadata sheet
        metadata_ws = wb.create_sheet("Metadata")
        metadata_ws.sheet_state = "hidden"
        
        # Write metadata
        self._write_metadata(metadata_ws, estimate, engagement, weeks, all_delivery_centers, filtered_roles, all_employees)
        
        # Write header rows
        try:
            self._write_headers(ws, estimate.phases, weeks, engagement.default_currency or "USD")
        except Exception as e:
            logger.error(f"Error writing headers: {e}", exc_info=True)
            raise ValueError(f"Failed to write headers: {str(e)}") from e
        
        # Write data rows WITH formulas BEFORE creating table
        # Excel Tables recognize formulas written before table creation as calculated columns
        # Ensure minimum of 20 data rows for easier row insertion
        min_rows = 20
        actual_num_rows = len(estimate.line_items)
        num_rows_to_write = max(actual_num_rows, min_rows)
        
        try:
            self._write_data_rows(ws, estimate.line_items, weeks, len(estimate.phases) if estimate.phases else 0, write_formulas=True, min_rows=min_rows)
        except Exception as e:
            logger.error(f"Error writing data rows: {e}", exc_info=True)
            raise ValueError(f"Failed to write data rows: {str(e)}") from e
        
        # Write totals row placeholder (just label, totals formulas written after table)
        try:
            self._write_totals_row(ws, num_rows_to_write, len(weeks))
        except Exception as e:
            logger.error(f"Error writing totals row: {e}", exc_info=True)
            raise ValueError(f"Failed to write totals row: {str(e)}") from e
        
        # Apply data validation and cell protection
        try:
            self._apply_validation_and_protection(
                ws, 
                all_delivery_centers, 
                filtered_roles, 
                all_employees,
                num_rows_to_write,
                len(weeks)
            )
        except Exception as e:
            logger.error(f"Error applying validation and protection: {e}", exc_info=True)
            raise ValueError(f"Failed to apply validation: {str(e)}") from e
        
        # Create Excel Table AFTER formulas are written
        # Excel will recognize the formulas as calculated columns
        # NOTE: Excel may remove merged cells in rows 1-2 during repair if they span table columns
        # This is a known Excel limitation - merged cells above tables can conflict
        try:
            self._create_excel_table(ws, num_rows_to_write, len(weeks))
        except Exception as e:
            logger.error(f"Error creating Excel table: {e}", exc_info=True)
            raise ValueError(f"Failed to create Excel table: {str(e)}") from e
        
        # No need to restore merged headers - we're using "visual merging" instead of actual merging
        # This avoids Excel repair warnings about merged cells conflicting with tables
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _generate_weeks_from_line_items(self, line_items: List[EstimateLineItem]) -> List[date]:
        """Generate list of week start dates from line items."""
        if not line_items:
            return []
        
        # Collect all week start dates from weekly hours
        week_dates = set()
        for line_item in line_items:
            if line_item.weekly_hours:
                for weekly_hour in line_item.weekly_hours:
                    week_dates.add(weekly_hour.week_start_date)
        
        # Also consider start/end dates to ensure coverage
        min_date = min(li.start_date for li in line_items)
        max_date = max(li.end_date for li in line_items)
        
        # Generate weeks between min and max dates
        weeks = []
        current = self._get_week_start(min_date)
        end_week_start = self._get_week_start(max_date)
        
        while current <= end_week_start:
            weeks.append(current)
            current += timedelta(days=7)
        
        # Add any weeks from weekly_hours that might be outside the date range
        for week_date in week_dates:
            if week_date not in weeks:
                weeks.append(week_date)
        
        return sorted(weeks)
    
    def _generate_weeks_from_dates(self, start_date: date, end_date: date) -> List[date]:
        """Generate list of week start dates between two dates."""
        weeks = []
        current = self._get_week_start(start_date)
        end_week_start = self._get_week_start(end_date)
        
        while current <= end_week_start:
            weeks.append(current)
            current += timedelta(days=7)
        
        return weeks
    
    def _get_week_start(self, d: date) -> date:
        """Get the Sunday (week start) for a given date."""
        days_since_sunday = (d.weekday() + 1) % 7
        return d - timedelta(days=days_since_sunday)
    
    def _safe_set_cell_value(self, ws, row: int, col: int, value):
        """Safely set a cell value, handling merged cells."""
        # Check if this cell is part of a merged range
        cell_coord = f"{get_column_letter(col)}{row}"
        for merged_range in ws.merged_cells.ranges:
            if cell_coord in str(merged_range):
                # Cell is merged - get the top-left cell of the merge
                top_left = merged_range.top_left
                if top_left[0] == row and top_left[1] == col:
                    # This IS the top-left cell, safe to set
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    return
                else:
                    # This is NOT the top-left cell, skip setting value
                    logger.warning(f"Cell {cell_coord} is part of merged range, skipping value assignment")
                    return
        
        # Cell is not merged, safe to set
        cell = ws.cell(row=row, column=col)
        cell.value = value
    
    def _write_metadata(self, ws, estimate: Estimate, engagement, weeks: List[date], 
                        delivery_centers: List[DeliveryCenter], roles: List[Role], employees: List[Employee]):
        """Write metadata to hidden sheet."""
        ws["A1"] = "estimate_id"
        ws["B1"] = str(estimate.id)
        ws["A2"] = "engagement_id"
        ws["B2"] = str(estimate.engagement_id)
        ws["A3"] = "engagement_delivery_center_id"
        ws["B3"] = str(engagement.delivery_center_id)
        ws["A4"] = "week_start_dates"
        ws["B4"] = ",".join([w.isoformat() for w in weeks])
        
        # Write phases
        if estimate.phases:
            ws["A5"] = "phases"
            for idx, phase in enumerate(estimate.phases):
                ws[f"B{5 + idx}"] = f"{phase.name}|{phase.start_date.isoformat()}|{phase.end_date.isoformat()}|{phase.color}"
        
        # Write delivery centers
        start_row = 10
        ws[f"A{start_row}"] = "delivery_centers"
        for idx, dc in enumerate(delivery_centers):
            ws[f"B{start_row + idx}"] = f"{dc.id}|{dc.name}"
        
        # Write roles
        roles_start = start_row + len(delivery_centers) + 2
        ws[f"A{roles_start}"] = "roles"
        for idx, role in enumerate(roles):
            ws[f"B{roles_start + idx}"] = f"{role.id}|{role.role_name}"
        
        # Write employees
        employees_start = roles_start + len(roles) + 2
        ws[f"A{employees_start}"] = "employees"
        for idx, emp in enumerate(employees):
            ws[f"B{employees_start + idx}"] = f"{emp.id}|{emp.first_name} {emp.last_name}"
    
    def _write_headers(self, ws, phases: Optional[List[EstimatePhase]], weeks: List[date], currency: str):
        """Write header rows to worksheet."""
        # Row 1: Phase headers (if phases exist)
        if phases and len(phases) > 0:
            col = 10  # Start after fixed columns (Payable Center, Role, Employee, Cost, Rate, Start Date, End Date, Billable, Billable %)
            phase_col_map = {}  # Map phase to column ranges
            
            # Collect all phase ranges and detect overlaps
            phase_ranges = []
            week_phases = {}  # Map week index to list of overlapping phases
            
            for phase in sorted(phases, key=lambda p: p.start_date):
                phase_start = self._get_week_start(phase.start_date)
                phase_end = self._get_week_start(phase.end_date)
                
                # Find columns for this phase
                start_col = None
                end_col = None
                for idx, week in enumerate(weeks):
                    # Check if week overlaps with phase (week is Sunday, phase dates are inclusive)
                    week_end = week + timedelta(days=6)  # Saturday
                    if week <= phase_end and week_end >= phase_start:
                        # Track which phases overlap this week
                        if idx not in week_phases:
                            week_phases[idx] = []
                        week_phases[idx].append(phase)
                        
                        if start_col is None:
                            start_col = col + idx
                        end_col = col + idx
                
                if start_col is not None and end_col is not None:
                    phase_ranges.append((phase, start_col, end_col))
            
            # FIRST: Unmerge ALL existing merges in row 1 before writing any phases
            # This prevents MergedCell issues when writing overlapping phases
            row_1_merges_to_remove = []
            for merged_range in list(ws.merged_cells.ranges):
                # Check if this merge is in row 1
                if merged_range.min_row <= 1 <= merged_range.max_row:
                    row_1_merges_to_remove.append(merged_range)
            
            for merged_range in row_1_merges_to_remove:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception as e:
                    logger.debug(f"Could not unmerge row 1 range {merged_range}: {e}")
            
            # NOW write phases - use "visual merging" instead of actual merging
            # Excel Tables cannot have merged cells that span table columns, so we simulate merging
            # by setting the same value and formatting in all cells of the range
            # For overlapping weeks, show both phase names
            
            # First, write all cells based on overlapping phases
            # This ensures overlapping weeks show all phases
            for week_idx, overlapping_phases_list in week_phases.items():
                col_idx = col + week_idx
                cell = ws.cell(row=1, column=col_idx)
                
                if len(overlapping_phases_list) > 1:
                    # Multiple phases overlap - show all phase names separated by " / "
                    phase_names = [p.name for p in overlapping_phases_list]
                    cell.value = " / ".join(phase_names)
                    
                    # Use diagonal stripe pattern to show both phase colors
                    # Get colors from both phases
                    first_phase = overlapping_phases_list[0]
                    second_phase = overlapping_phases_list[1] if len(overlapping_phases_list) > 1 else first_phase
                    
                    color1_hex = first_phase.color.replace("#", "")
                    color2_hex = second_phase.color.replace("#", "")
                    
                    # Use diagonal stripe pattern to show both colors
                    if len(color1_hex) == 6 and len(color2_hex) == 6:
                        # Diagonal stripe pattern - shows both colors
                        cell.fill = PatternFill(
                            start_color=color1_hex,
                            end_color=color2_hex,
                            fill_type="darkUp"  # Diagonal stripes
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif len(color1_hex) == 6:
                        # Fallback to first phase color if second doesn't have valid color
                        cell.fill = PatternFill(start_color=color1_hex, end_color=color1_hex, fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # Single phase - will be handled in next loop
                    pass
            
            # Then, write single-phase cells (non-overlapping)
            for phase, start_col, end_col in phase_ranges:
                try:
                    # Handle color - remove # if present and ensure 6 hex digits
                    color_hex = phase.color.replace("#", "")
                    fill_color = color_hex if len(color_hex) == 6 else None
                    
                    # Set value and formatting in ALL cells of the range (not just top-left)
                    # Skip cells that already have overlapping phases written
                    for col_idx in range(start_col, end_col + 1):
                        week_idx = col_idx - col  # Convert column to week index
                        cell = ws.cell(row=1, column=col_idx)
                        
                        # Only write if this cell doesn't already have overlapping phases
                        if week_idx not in week_phases or len(week_phases[week_idx]) == 1:
                            cell.value = phase.name
                            if fill_color:
                                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                            else:
                                cell.font = Font(bold=True)
                            
                            # Center alignment only in leftmost cell, left align others for visual consistency
                            if col_idx == start_col:
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            else:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    logger.debug(f"Created visual merge for phase '{phase.name}' (columns {start_col}-{end_col})")
                except Exception as e:
                    logger.error(f"Error writing phase '{phase.name}' header: {e}", exc_info=True)
                    # Continue with next phase
                phase_col_map[phase.id] = (start_col, end_col)
        
        # Row 2: Year headers
        col = 10
        current_year = None
        year_start_col = None
        year_ranges = []  # Store year ranges to merge after collecting all
        
        for idx, week in enumerate(weeks):
            if week.year != current_year:
                if current_year is not None and year_start_col is not None:
                    # Store range for previous year
                    year_ranges.append((year_start_col, col + idx - 1, current_year))
                current_year = week.year
                year_start_col = col + idx
        
        # Store last year range
        if current_year is not None and year_start_col is not None:
            year_ranges.append((year_start_col, col + len(weeks) - 1, current_year))
        
        # Now set values for all year ranges using "visual merging" instead of actual merging
        # Excel Tables cannot have merged cells that span table columns
        for start_col, end_col, year in year_ranges:
            # Set value and formatting in ALL cells of the range (not just top-left)
            # This creates a "visually merged" appearance without actual merging
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=2, column=col)
                cell.value = year  # Same value in all cells
                cell.font = Font(bold=True)
                # Center alignment only in leftmost cell, left align others for visual consistency
                if col == start_col:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Row 3: Week headers
        col = 10
        for idx, week in enumerate(weeks):
            cell = ws.cell(row=3, column=col + idx)
            cell.value = week.strftime("%m/%d/%Y")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.protection = Protection(locked=True)
        
        # Row 4: Column headers
        headers = [
            "Payable Center",
            "Role",
            "Employee",
            f"Cost ({currency})",
            f"Rate ({currency})",
            "Start Date",
            "End Date",
            "Billable",
            "Billable %",
        ]
        
        for idx, header in enumerate(headers):
            cell = ws[f"{get_column_letter(idx + 1)}4"]
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Week column headers
        # IMPORTANT: Excel Tables require unique column headers
        # So we need to give each week column a unique header
        for idx, week in enumerate(weeks):
            cell = ws[f"{get_column_letter(col + idx)}4"]
            # Use week date as part of header to make it unique
            # Format: "Hours_MMDD" or just "Hours" with index for uniqueness
            cell.value = f"Hours{idx+1}"  # Unique header: Hours1, Hours2, etc.
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.protection = Protection(locked=True)
        
        # Totals column headers
        totals_start_col = col + len(weeks)
        total_headers = [
            "Total Hours",
            "Total Cost",
            "Total Revenue",
            "Billable Expense Amount",
            "Margin Amount",
            "Margin % (Without Expenses)",
            "Margin % (With Expenses)",
        ]
        
        for idx, header in enumerate(total_headers):
            cell = ws[f"{get_column_letter(totals_start_col + idx)}4"]
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.protection = Protection(locked=True)
    
    def _write_data_rows(self, ws, line_items: List[EstimateLineItem], weeks: List[date], phase_rows: int, write_formulas: bool = True, min_rows: int = 20):
        """Write data rows to worksheet with formulas written BEFORE table creation.
        
        Args:
            ws: Worksheet to write to
            line_items: List of line items to write
            weeks: List of week dates
            phase_rows: Number of phase header rows (unused, kept for compatibility)
            write_formulas: Whether to write calculated column formulas
            min_rows: Minimum number of data rows to create (default 20)
        """
        start_row = 5  # After header rows
        week_col_start = 10  # Column J
        totals_start_col = week_col_start + len(weeks)
        header_row = 4
        
        # Ensure we write at least min_rows rows
        num_rows_to_write = max(len(line_items), min_rows)
        
        # Read header names and build formulas ONCE (before the loop)
        if write_formulas:
            cost_header = str(ws.cell(row=header_row, column=4).value or "Cost")
            rate_header = str(ws.cell(row=header_row, column=5).value or "Rate")
            billable_pct_header = str(ws.cell(row=header_row, column=9).value or "Billable %")
            total_hours_header = str(ws.cell(row=header_row, column=totals_start_col).value or "Total Hours")
            total_cost_header = str(ws.cell(row=header_row, column=totals_start_col + 1).value or "Total Cost")
            total_revenue_header = str(ws.cell(row=header_row, column=totals_start_col + 2).value or "Total Revenue")
            billable_expense_header = str(ws.cell(row=header_row, column=totals_start_col + 3).value or "Billable Expense Amount")
            margin_amount_header = str(ws.cell(row=header_row, column=totals_start_col + 4).value or "Margin Amount")
            
            # Escape column names (only parentheses and % need escaping)
            def escape_col_name(name: str) -> str:
                if '(' in name or ')' in name or '%' in name:
                    return f"[{name}]"
                return name
            
            cost_ref = escape_col_name(cost_header)
            rate_ref = escape_col_name(rate_header)
            billable_pct_ref = escape_col_name(billable_pct_header)
            
            # Build week sum reference
            week_col_refs = []
            for week_idx in range(len(weeks)):
                week_col_refs.append(f"[@Hours{week_idx+1}]")
            week_sum_ref = "+".join(week_col_refs)
        
        # Write line items and empty rows (minimum min_rows rows total)
        for row_idx in range(num_rows_to_write):
            row = start_row + row_idx
            line_item = line_items[row_idx] if row_idx < len(line_items) else None
            
            # If no line item for this row, write empty row with formulas only
            if line_item is None:
                # Write formulas for empty rows so they're ready when user adds data
                if write_formulas:
                    totals_start_col = week_col_start + len(weeks)
                    first_week_col = get_column_letter(week_col_start)
                    last_week_col = get_column_letter(week_col_start + len(weeks) - 1)
                    ws.cell(row=row, column=totals_start_col).value = f"=SUM({first_week_col}{row}:{last_week_col}{row})"
                    ws.cell(row=row, column=totals_start_col + 1).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(4)}{row}"
                    ws.cell(row=row, column=totals_start_col + 2).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(5)}{row}"
                    ws.cell(row=row, column=totals_start_col + 3).value = f"={get_column_letter(totals_start_col + 2)}{row}*{get_column_letter(9)}{row}/100"
                    ws.cell(row=row, column=totals_start_col + 4).value = f"={get_column_letter(totals_start_col + 2)}{row}-{get_column_letter(totals_start_col + 1)}{row}"
                    ws.cell(row=row, column=totals_start_col + 5).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,({get_column_letter(totals_start_col + 4)}{row}/{get_column_letter(totals_start_col + 2)}{row})*100)"
                    ws.cell(row=row, column=totals_start_col + 6).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,(({get_column_letter(totals_start_col + 4)}{row}-{get_column_letter(totals_start_col + 3)}{row})/{get_column_letter(totals_start_col + 2)}{row})*100)"
                continue
            
            # Write line item data
            # Payable Center
            if line_item.role_rate and line_item.role_rate.delivery_center:
                ws.cell(row=row, column=1).value = line_item.role_rate.delivery_center.name
            
            # Role
            if line_item.role_rate and line_item.role_rate.role:
                ws.cell(row=row, column=2).value = line_item.role_rate.role.role_name
            
            # Employee
            if line_item.employee:
                ws.cell(row=row, column=3).value = f"{line_item.employee.first_name} {line_item.employee.last_name}"
            
            # Cost
            ws.cell(row=row, column=4).value = float(line_item.cost)
            
            # Rate
            ws.cell(row=row, column=5).value = float(line_item.rate)
            
            # Start Date
            ws.cell(row=row, column=6).value = line_item.start_date
            
            # End Date
            ws.cell(row=row, column=7).value = line_item.end_date
            
            # Billable
            ws.cell(row=row, column=8).value = "Yes" if line_item.billable else "No"
            
            # Billable %
            ws.cell(row=row, column=9).value = float(line_item.billable_expense_percentage)
            
            # Weekly hours
            weekly_hours_dict = {}
            if line_item.weekly_hours:
                for wh in line_item.weekly_hours:
                    weekly_hours_dict[wh.week_start_date] = float(wh.hours)
            
            for week_idx, week in enumerate(weeks):
                col = week_col_start + week_idx
                hours = weekly_hours_dict.get(week, 0)
                ws.cell(row=row, column=col).value = hours
            
            # Write calculated column formulas BEFORE table creation
            # Use regular cell references - Excel will convert them to structured references when table is created
            if write_formulas:
                totals_start_col = week_col_start + len(weeks)
                
                # Total Hours: SUM of all week columns using regular cell references
                first_week_col = get_column_letter(week_col_start)
                last_week_col = get_column_letter(week_col_start + len(weeks) - 1)
                ws.cell(row=row, column=totals_start_col).value = f"=SUM({first_week_col}{row}:{last_week_col}{row})"
                
                # Total Cost: Total Hours * Cost
                ws.cell(row=row, column=totals_start_col + 1).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(4)}{row}"
                
                # Total Revenue: Total Hours * Rate
                ws.cell(row=row, column=totals_start_col + 2).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(5)}{row}"
                
                # Billable Expense Amount: Total Revenue * Billable % / 100
                ws.cell(row=row, column=totals_start_col + 3).value = f"={get_column_letter(totals_start_col + 2)}{row}*{get_column_letter(9)}{row}/100"
                
                # Margin Amount: Total Revenue - Total Cost
                ws.cell(row=row, column=totals_start_col + 4).value = f"={get_column_letter(totals_start_col + 2)}{row}-{get_column_letter(totals_start_col + 1)}{row}"
                
                # Margin % Without Expenses
                ws.cell(row=row, column=totals_start_col + 5).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,({get_column_letter(totals_start_col + 4)}{row}/{get_column_letter(totals_start_col + 2)}{row})*100)"
                
                # Margin % With Expenses
                ws.cell(row=row, column=totals_start_col + 6).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,(({get_column_letter(totals_start_col + 4)}{row}-{get_column_letter(totals_start_col + 3)}{row})/{get_column_letter(totals_start_col + 2)}{row})*100)"
    
    def _escape_column_name(self, col_name: str) -> str:
        """Escape column name for use in structured references.
        
        Excel handles special characters automatically, but we need to ensure
        the name matches exactly what Excel sees in the table.
        For names with special characters, Excel may sanitize them.
        We'll use the name as-is - Excel will handle escaping.
        """
        # Excel handles special characters automatically in structured references
        # Just return the name as-is - it will work in [@ColumnName] format
        return col_name
    
    def _write_calculated_column_formulas(self, ws, line_items: List[EstimateLineItem], weeks: List[date]):
        """Write calculated column formulas AFTER table is created.
        
        This ensures structured references resolve correctly.
        Column names with special characters must be properly escaped.
        """
        if not line_items:
            return
        
        start_row = 5
        week_col_start = 10
        totals_start_col = week_col_start + len(weeks)
        header_row = 4
        
        # Read EXACT header names from row 4 (table now exists, so these are the actual column names)
        # Get cell values directly and ensure they're strings
        cost_header = str(ws.cell(row=header_row, column=4).value or "").strip()
        rate_header = str(ws.cell(row=header_row, column=5).value or "").strip()
        billable_pct_header = str(ws.cell(row=header_row, column=9).value or "").strip()
        
        # Read totals column headers - these MUST exist
        total_hours_header = str(ws.cell(row=header_row, column=totals_start_col).value or "").strip()
        total_cost_header = str(ws.cell(row=header_row, column=totals_start_col + 1).value or "").strip()
        total_revenue_header = str(ws.cell(row=header_row, column=totals_start_col + 2).value or "").strip()
        billable_expense_header = str(ws.cell(row=header_row, column=totals_start_col + 3).value or "").strip()
        margin_amount_header = str(ws.cell(row=header_row, column=totals_start_col + 4).value or "").strip()
        
        # Log the actual column names for debugging
        logger.info(f"Column names read from headers:")
        logger.info(f"  Cost: '{cost_header}'")
        logger.info(f"  Rate: '{rate_header}'")
        logger.info(f"  Billable %: '{billable_pct_header}'")
        logger.info(f"  Total Hours: '{total_hours_header}'")
        logger.info(f"  Total Cost: '{total_cost_header}'")
        logger.info(f"  Total Revenue: '{total_revenue_header}'")
        
        # Build structured reference for week columns
        week_col_refs = []
        for week_idx in range(len(weeks)):
            week_col_refs.append(f"[@Hours{week_idx+1}]")
        week_sum_ref = "+".join(week_col_refs)
        
        # Build formulas using structured references
        # CRITICAL: Excel handles spaces automatically in structured references
        # Only parentheses and % need special escaping: [@[Column (Name)]] or [@[Column%]]
        # Spaces are fine: [@Column Name] works without extra brackets
        def escape_col_name(name: str) -> str:
            """Escape column name for structured reference if it has special characters.
            
            Excel structured reference format:
            - Simple name or name with spaces: [@Column Name] (Excel handles spaces automatically)
            - Name with parentheses or %: [@[Column (Name)]] (needs inner brackets)
            """
            # Only escape if name has parentheses or % - Excel handles spaces automatically
            if '(' in name or ')' in name or '%' in name:
                return f"[{name}]"
            return name
        
        # Escape column names for structured references
        # Only escape names with parentheses or % - Excel handles spaces automatically
        cost_ref = escape_col_name(cost_header)
        rate_ref = escape_col_name(rate_header)
        billable_pct_ref = escape_col_name(billable_pct_header)
        # Totals columns don't need escaping if they only have spaces
        total_hours_ref = total_hours_header  # "Total Hours" - space is fine
        total_cost_ref = total_cost_header    # "Total Cost" - space is fine
        total_revenue_ref = total_revenue_header  # "Total Revenue" - space is fine
        billable_expense_ref = billable_expense_header  # "Billable Expense Amount" - space is fine
        margin_amount_ref = margin_amount_header  # "Margin Amount" - space is fine
        
        # Build formulas in dependency order
        # Use the exact column names - Excel will handle spaces automatically in [@Column Name] format
        # Only use [@[Column Name]] format for names with parentheses or %
        formulas = {
            totals_start_col: f"={week_sum_ref}",  # Total Hours - no dependencies
            totals_start_col + 1: f"=[@{total_hours_ref}]*[@{cost_ref}]",  # Total Cost
            totals_start_col + 2: f"=[@{total_hours_ref}]*[@{rate_ref}]",  # Total Revenue
            totals_start_col + 3: f"=[@{total_revenue_ref}]*[@{billable_pct_ref}]/100",  # Billable Expense Amount
            totals_start_col + 4: f"=[@{total_revenue_ref}]-[@{total_cost_ref}]",  # Margin Amount
            totals_start_col + 5: f"=IF([@{total_revenue_ref}]=0,0,([@{margin_amount_ref}]/[@{total_revenue_ref}])*100)",  # Margin % Without Expenses
            totals_start_col + 6: f"=IF([@{total_revenue_ref}]=0,0,(([@{margin_amount_ref}]-[@{billable_expense_ref}])/[@{total_revenue_ref}])*100)",  # Margin % With Expenses
        }
        
        # Log escaped column names for debugging
        logger.info("Escaped column names for structured references:")
        logger.info(f"  Cost: '{cost_ref}' (from '{cost_header}')")
        logger.info(f"  Rate: '{rate_ref}' (from '{rate_header}')")
        logger.info(f"  Billable %: '{billable_pct_ref}' (from '{billable_pct_header}')")
        logger.info(f"  Total Hours: '{total_hours_ref}' (from '{total_hours_header}')")
        logger.info(f"  Total Cost: '{total_cost_ref}' (from '{total_cost_header}')")
        logger.info(f"  Total Revenue: '{total_revenue_ref}' (from '{total_revenue_header}')")
        logger.info(f"  Billable Expense: '{billable_expense_ref}' (from '{billable_expense_header}')")
        logger.info(f"  Margin Amount: '{margin_amount_ref}' (from '{margin_amount_header}')")
        
        # Log formulas for debugging
        logger.info("Calculated column formulas:")
        for col, formula in formulas.items():
            logger.info(f"  Column {col}: {formula}")
        
        # Write formulas to all data rows
        try:
            for row_idx in range(len(line_items)):
                row = start_row + row_idx
                for col, formula in formulas.items():
                    try:
                        cell = ws.cell(row=row, column=col)
                        cell.value = formula
                        logger.debug(f"Wrote formula to row {row}, col {col}: {formula}")
                    except Exception as e:
                        logger.error(f"Error writing formula to row {row}, col {col}: {e}", exc_info=True)
                        raise
            logger.info(f"Successfully wrote calculated column formulas to {len(line_items)} data rows")
        except Exception as e:
            logger.error(f"Failed to write calculated column formulas: {e}", exc_info=True)
            raise ValueError(f"Failed to write calculated column formulas: {str(e)}") from e
    
    def _write_totals_row(self, ws, num_rows: int, num_weeks: int):
        """Write totals row placeholder - formulas will be written AFTER table is created.
        
        We only write the "TOTALS" label here. The actual totals formulas must be
        written AFTER the table is created so structured references resolve correctly.
        """
        if num_rows == 0:
            return
        
        header_row = 4
        start_row = 5
        end_row = start_row + num_rows - 1
        totals_row = end_row + 1  # After data rows (this will be INSIDE the table)
        
        # Set "TOTALS" label in column A only
        # Formulas will be written after table creation in _create_excel_table
        ws.cell(row=totals_row, column=1).value = "TOTALS"
        ws.cell(row=totals_row, column=1).font = Font(bold=True)
    
    def _apply_validation_and_protection(self, ws, delivery_centers: List[DeliveryCenter], 
                                        roles: List[Role], employees: List[Employee],
                                        num_rows: int, num_weeks: int):
        """Apply data validation and cell protection."""
        start_row = 5
        end_row = 4 + num_rows
        
        # Calculate max row for validation
        # Since we're using Excel Tables, validation will automatically extend when rows are inserted
        # Set validation to cover data rows and totals row only (totals row is the last row)
        totals_row = start_row + num_rows  # Totals row is the last row
        max_validation_row = totals_row  # Only up to totals row
        
        # Payable Center dropdown (Column A)
        dc_names = [dc.name for dc in delivery_centers]
        if dc_names:
            dv = DataValidation(type="list", formula1=f'"{",".join(dc_names)}"', allow_blank=True)
            dv.add(f"A{start_row}:A{max_validation_row}")
            ws.add_data_validation(dv)
        
        # Role dropdown (Column B) - filtered by engagement delivery center
        role_names = [role.role_name for role in roles]
        if role_names:
            dv = DataValidation(type="list", formula1=f'"{",".join(role_names)}"', allow_blank=True)
            dv.add(f"B{start_row}:B{max_validation_row}")
            ws.add_data_validation(dv)
        
        # Employee dropdown (Column C)
        emp_names = [f"{emp.first_name} {emp.last_name}" for emp in employees]
        if emp_names:
            dv = DataValidation(type="list", formula1=f'"{",".join(emp_names)}"', allow_blank=True)
            dv.add(f"C{start_row}:C{max_validation_row}")
            ws.add_data_validation(dv)
        
        # Date validation for Start Date (Column F) and End Date (Column G)
        date_dv = DataValidation(type="date", operator="between", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
        date_dv.add(f"F{start_row}:F{max_validation_row}")
        date_dv.add(f"G{start_row}:G{max_validation_row}")
        ws.add_data_validation(date_dv)
        
        # Number validation for Cost (Column D), Rate (Column E), Billable % (Column I)
        number_dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        number_dv.add(f"D{start_row}:D{max_validation_row}")
        number_dv.add(f"E{start_row}:E{max_validation_row}")
        number_dv.add(f"I{start_row}:I{max_validation_row}")
        ws.add_data_validation(number_dv)
        
        # Billable % should be 0-100
        pct_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="100", allow_blank=True)
        pct_dv.add(f"I{start_row}:I{max_validation_row}")
        ws.add_data_validation(pct_dv)
        
        # Hours validation (week columns)
        week_col_start = 10
        hours_dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        for week_idx in range(num_weeks):
            col = week_col_start + week_idx
            col_letter = get_column_letter(col)
            hours_dv.add(f"{col_letter}{start_row}:{col_letter}{max_validation_row}")
        ws.add_data_validation(hours_dv)
        
        # Calculate max rows and columns to unlock
        # Only unlock cells in actual data rows and totals row, not empty rows below
        # Excel Table will handle protection for new rows when they're inserted
        totals_row = start_row + num_rows  # Totals row
        max_row = totals_row  # Only up to totals row (the last row)
        max_col = week_col_start + num_weeks + 7  # Fixed columns + weeks + totals columns
        
        # Create protection objects to reuse
        unlocked_protection = Protection(locked=False)
        locked_protection = Protection(locked=True)
        
        # FIRST: Unlock ALL cells explicitly (cells are locked by default in Excel)
        # This must happen before we lock specific cells
        logger.info(f"Unlocking all cells in worksheet (rows 1-{max_row}, cols 1-{max_col})")
        
        # Unlock ALL cells in the range - iterate through all possible cells
        # This ensures cells are unlocked even if they don't have values yet
        unlocked_count = 0
        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row_num, column=col_num)
                    # Set protection to unlocked using both methods to ensure it works
                    cell.protection = unlocked_protection
                    # Also set the locked attribute directly as a backup
                    if hasattr(cell, 'protection') and hasattr(cell.protection, 'locked'):
                        cell.protection.locked = False
                    unlocked_count += 1
                except Exception as e:
                    logger.debug(f"Could not unlock cell row={row_num}, col={col_num}: {e}")
        
        logger.info(f"Unlocked {unlocked_count} cells")
        
        # NOW lock only specific cells that should be protected
        
        # Lock header rows (rows 1-4) - users cannot modify headers
        logger.info("Locking header rows 1-4")
        for row in range(1, 5):
            for col in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.protection = locked_protection
                except Exception:
                    pass
        
        # Lock week header columns (row 3) - users cannot modify week headers
        logger.info(f"Locking week header columns in row 3")
        for week_idx in range(num_weeks):
            col = week_col_start + week_idx
            try:
                cell = ws.cell(row=3, column=col)
                cell.protection = locked_protection
            except Exception:
                pass
        
        # Lock totals columns (formulas) - users cannot modify calculated totals
        totals_start_col = week_col_start + num_weeks
        logger.info(f"Locking totals columns {totals_start_col} to {totals_start_col + 6}")
        for col_offset in range(7):
            col = totals_start_col + col_offset
            for row in range(start_row, max_row + 1):  # Lock for all rows including new ones
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.protection = locked_protection
                except Exception:
                    pass
        
        # Lock totals row formulas (if it exists)
        if num_rows > 0:
            totals_row = 5 + num_rows
            logger.info(f"Locking totals row {totals_row}")
            # Lock week column totals in totals row
            for week_idx in range(num_weeks):
                col = week_col_start + week_idx
                try:
                    cell = ws.cell(row=totals_row, column=col)
                    cell.protection = locked_protection
                except Exception:
                    pass
            # Lock totals column totals in totals row
            for col_offset in range(7):
                col = totals_start_col + col_offset
                try:
                    cell = ws.cell(row=totals_row, column=col)
                    cell.protection = locked_protection
                except Exception:
                    pass
        
        # Enable sheet protection but allow editing unlocked cells
        # IMPORTANT: Configure protection settings BEFORE enabling sheet protection
        logger.info("Configuring sheet protection")
        
        # Configure all protection settings first
        ws.protection.password = ""  # Empty string for no password
        ws.protection.formatCells = True  # Allow formatting cells
        ws.protection.formatColumns = True
        ws.protection.formatRows = True
        ws.protection.insertColumns = False
        ws.protection.insertRows = True  # Allow inserting rows
        ws.protection.insertHyperlinks = True
        ws.protection.deleteColumns = False
        ws.protection.deleteRows = True  # Allow deleting rows
        ws.protection.selectLockedCells = True  # Can select locked cells
        ws.protection.selectUnlockedCells = True  # Can select unlocked cells
        ws.protection.sort = True
        ws.protection.autoFilter = True
        ws.protection.pivotTables = True
        
        # Final unlock pass - ensure all editable cells are unlocked BEFORE enabling protection
        logger.info("Final unlock pass - ensuring editable cells are unlocked")
        try:
            # Unlock all editable data cells (columns A-I and week columns)
            for row_num in range(start_row, max_row + 1):
                # Unlock editable columns: A-I (Payable Center through Billable %)
                for col_num in range(1, 10):
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        # Force unlock - create new Protection object
                        cell.protection = Protection(locked=False)
                        cell.protection.locked = False
                    except Exception:
                        pass
                
                # Unlock week columns (starting at column 10)
                for week_idx in range(num_weeks):
                    col = week_col_start + week_idx
                    try:
                        cell = ws.cell(row=row_num, column=col)
                        # Force unlock - create new Protection object
                        cell.protection = Protection(locked=False)
                        cell.protection.locked = False
                    except Exception:
                        pass
            
            # Verify sample cells are unlocked
            sample_cell_a = ws.cell(row=start_row, column=1)
            sample_cell_d = ws.cell(row=start_row, column=4)
            sample_cell_j = ws.cell(row=start_row, column=10)
            logger.info(f"Before protection - A{start_row} locked: {sample_cell_a.protection.locked}")
            logger.info(f"Before protection - D{start_row} locked: {sample_cell_d.protection.locked}")
            logger.info(f"Before protection - J{start_row} locked: {sample_cell_j.protection.locked}")
        except Exception as e:
            logger.warning(f"Could not perform final unlock: {e}")
        
        # DO NOT enable sheet protection - it causes all cells to be locked
        # Instead, rely on data validation to enforce rules
        # Note: Without sheet protection, formula cells won't be protected from editing
        # but data validation will still enforce dropdowns and value constraints
        ws.protection.sheet = False
        logger.info("Sheet protection DISABLED - all cells are editable")
        logger.warning("NOTE: Totals formula cells are NOT protected from editing. Users should not modify formulas.")
    
    def _create_excel_table(self, ws, num_rows: int, num_weeks: int):
        """Create Excel Table with built-in totals row for automatic formula copying.
        
        The table includes:
        - Row 4: Headers
        - Rows 5 to (5 + num_rows - 1): Data rows
        - Excel will add a totals row automatically when enabled
        
        When rows are inserted, Excel will automatically:
        1. Copy formulas from the row above
        2. Extend the table to include the new row
        3. Update the totals row to include all data rows
        """
        if num_rows == 0:
            return
        
        header_row = 4  # Column headers are in row 4
        start_row = 5  # Data starts at row 5
        end_row = start_row + num_rows - 1  # Last data row
        totals_row = end_row + 1  # Totals row (MUST be INSIDE the table)
        
        # Calculate column ranges
        # Fixed columns: A-I (Payable Center through Billable %)
        # Week columns: J onwards (10 + num_weeks)
        # Totals columns: after week columns
        week_col_start = 10
        totals_start_col = week_col_start + num_weeks
        last_col = totals_start_col + 6  # 7 totals columns (column AM)
        
        # Verify all header cells have values (Excel Tables require this)
        # Check row 4 headers to ensure they're valid
        for col in range(1, last_col + 1):
            try:
                cell = ws.cell(row=header_row, column=col)
                if cell.value is None or str(cell.value).strip() == "":
                    # Set a default header if empty
                    cell.value = f"Column{col}"
                    logger.warning(f"Empty header at row {header_row}, col {col}, set to 'Column{col}'")
            except Exception as e:
                logger.warning(f"Could not check header cell row {header_row}, col {col}: {e}")
        
        # OPTION A: Table range INCLUDES the totals row
        # Table ref: A4:AM7 (header + data rows + totals row)
        # totalsRowCount=1 means the LAST row (row 7) is the totals row
        table_ref = f"A{header_row}:{get_column_letter(last_col)}{totals_row}"
        
        # Verify the range is valid and doesn't include merged cells in header row
        # Excel Tables cannot have merged cells in the header row
        # Also check that no merged cells overlap with the table range
        try:
            # Check if row 4 (header row) has any merged cells (it shouldn't)
            merged_in_header = False
            merged_overlapping_table = False
            for merged_range in ws.merged_cells.ranges:
                # Check if merged cells are in header row
                if merged_range.min_row <= header_row <= merged_range.max_row:
                    merged_in_header = True
                    logger.warning(f"Merged cells found in header row: {merged_range}")
                # Check if merged cells overlap with table range at all
                # Table range is from header_row to totals_row, columns 1 to last_col
                if (merged_range.min_row <= totals_row and merged_range.max_row >= header_row and
                    merged_range.min_col <= last_col and merged_range.max_col >= 1):
                    # Only flag if it's actually overlapping the data/totals area (not just phase headers above)
                    if merged_range.max_row >= start_row:
                        merged_overlapping_table = True
                        logger.warning(f"Merged cells overlapping table range: {merged_range}")
            
            if merged_in_header:
                logger.error("Excel Table cannot be created: header row contains merged cells")
                raise ValueError("Cannot create Excel Table: header row (row 4) contains merged cells")
            
            if merged_overlapping_table:
                logger.warning("Merged cells overlap table range - this may cause Excel repair warnings")
            
            # Create table with OPTION A: Totals row INSIDE the table
            table = Table(displayName="EstimateData", ref=table_ref)
            
            # Configure table structure
            table.headerRowCount = 1  # First row (row 4) is header
            table.totalsRowCount = 1  # Last row (totals_row) is totals row
            
            # Set autoFilter range - MUST exclude totals row
            # Filter should cover header + data rows only (A4:AM6), not totals row
            auto_filter_ref = f"A{header_row}:{get_column_letter(last_col)}{end_row}"
            table.autoFilter = AutoFilter(ref=auto_filter_ref)
            
            # Set table style
            style = TableStyleInfo(
                name="TableStyleLight9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            # Ensure totals row cells are empty (not None, not empty string) before creating table
            # Excel Tables work best when totals row cells are truly empty until formulas are written
            # Only ensure column A has "TOTALS" label - other cells should be None
            for col in range(2, last_col + 1):  # Skip column A (has "TOTALS" label)
                try:
                    cell = ws.cell(row=totals_row, column=col)
                    # Ensure totals row cells (except column A) are None, not empty string
                    if cell.value == "":
                        cell.value = None
                except Exception:
                    pass
            
            # Add table to worksheet
            ws.add_table(table)
            
            # NOW write totals row formulas AFTER table exists
            
            # Totals for week columns (sum of hours)
            # Use SUBTOTAL with column ranges - Excel Tables recognize this as totals row formulas
            for week_idx in range(num_weeks):
                col = week_col_start + week_idx
                col_letter = get_column_letter(col)
                cell = ws.cell(row=totals_row, column=col)
                # SUBTOTAL(109, ...) is SUM that ignores hidden rows and excludes totals row
                # Reference the data rows only (start_row to end_row)
                cell.value = f"=SUBTOTAL(109,{col_letter}{start_row}:{col_letter}{end_row})"
                cell.font = Font(bold=True)
            
            # Totals for totals columns (sum of calculated totals from each row)
            # Use SUBTOTAL with column ranges - Excel Tables recognize this as totals row formulas
            for col_offset in range(7):  # 7 totals columns
                col = totals_start_col + col_offset
                col_letter = get_column_letter(col)
                cell = ws.cell(row=totals_row, column=col)
                # SUBTOTAL(109, ...) is SUM that ignores hidden rows and excludes totals row
                # Reference the data rows only (start_row to end_row)
                cell.value = f"=SUBTOTAL(109,{col_letter}{start_row}:{col_letter}{end_row})"
                cell.font = Font(bold=True)
            
            logger.info(f"Created Excel Table 'EstimateData' with range {table_ref}")
            logger.info(f"Table includes totals row at row {totals_row} (INSIDE table)")
            logger.info(f"AutoFilter range: {auto_filter_ref} (excludes totals row)")
            logger.info(f"Data rows: {start_row} to {end_row}, Totals row: {totals_row}")
            logger.info("Totals formulas use SUBTOTAL - should work correctly in Excel Table")
            
            # Note: We use "visual merging" (same value/formatting in multiple cells) instead of actual merging
            # This avoids Excel repair warnings about merged cells conflicting with tables
            #
            # Excel may still show a repair warning about the Table being repaired.
            # This is expected when formulas are written before the table exists - Excel needs to
            # convert regular cell references to structured references. The repair is harmless
            # and all formulas will work correctly after Excel processes the file.
        except Exception as e:
            logger.error(f"Failed to create Excel Table: {e}", exc_info=True)
            raise ValueError(f"Failed to create Excel Table: {str(e)}") from e

