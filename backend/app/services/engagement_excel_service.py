"""
Excel export/import service for engagements (Resource Plan).
"""

import logging
import io
from typing import List, Dict, Optional, Tuple
from uuid import UUID
from datetime import date, datetime, timedelta
from decimal import Decimal
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter

from app.db.repositories.engagement_repository import EngagementRepository
from app.db.repositories.engagement_line_item_repository import EngagementLineItemRepository
from app.db.repositories.engagement_weekly_hours_repository import EngagementWeeklyHoursRepository
from app.db.repositories.delivery_center_repository import DeliveryCenterRepository
from app.db.repositories.employee_repository import EmployeeRepository
from app.db.repositories.role_repository import RoleRepository
from app.db.repositories.role_rate_repository import RoleRateRepository
from app.db.repositories.opportunity_repository import OpportunityRepository
from app.models.engagement import Engagement, EngagementLineItem, EngagementWeeklyHours, EngagementPhase
from app.models.role import Role
from app.utils.currency_converter import convert_currency

logger = logging.getLogger(__name__)


class EngagementExcelService:
    """Service for exporting/importing engagement Resource Plans to/from Excel."""
    
    def __init__(self, session: AsyncSession):
        self.session = session
        self.engagement_repo = EngagementRepository(session)
        self.line_item_repo = EngagementLineItemRepository(session)
        self.weekly_hours_repo = EngagementWeeklyHoursRepository(session)
        self.delivery_center_repo = DeliveryCenterRepository(session)
        self.employee_repo = EmployeeRepository(session)
        self.role_repo = RoleRepository(session)
        self.role_rate_repo = RoleRateRepository(session)
        self.opportunity_repo = OpportunityRepository(session)
        # Store model classes for use in queries
        from app.models.role import Role
        from app.models.role_rate import RoleRate
        self.role_model = Role
        self.role_rate_model = RoleRate
    
    async def export_engagement_to_excel(self, engagement_id: UUID) -> io.BytesIO:
        """Export an engagement Resource Plan to Excel format.
        
        This mirrors the Estimate Excel export but for Resource Plans.
        Dates are flexible (not tied to Opportunity dates).
        """
        # Get engagement with all relationships
        engagement = await self.engagement_repo.get_with_line_items(engagement_id)
        if not engagement:
            raise ValueError("Engagement not found")
        
        # Ensure line items are sorted by row_order
        if engagement.line_items:
            engagement.line_items = sorted(engagement.line_items, key=lambda li: li.row_order if li.row_order is not None else 0)
        
        # Get opportunity for delivery center and currency
        opportunity = await self.opportunity_repo.get(engagement.opportunity_id)
        if not opportunity:
            raise ValueError("Opportunity not found")
        
        opportunity_delivery_center_id = opportunity.delivery_center_id
        if not opportunity_delivery_center_id:
            raise ValueError("Opportunity Invoice Center (delivery_center_id) is required")
        
        # Get all delivery centers, employees, and roles for dropdowns
        all_delivery_centers = await self.delivery_center_repo.list_all()
        all_employees = await self.employee_repo.list(skip=0, limit=10000)
        
        # Get roles filtered by opportunity delivery center
        from app.models.role_rate import RoleRate
        roles_result = await self.session.execute(
            select(Role)
            .join(RoleRate, Role.id == RoleRate.role_id)
            .where(RoleRate.delivery_center_id == opportunity_delivery_center_id)
            .distinct()
        )
        filtered_roles = list(roles_result.scalars().all())
        
        # Generate weeks from earliest Start Date week and latest End Date week of any line item (resource plan role)
        if engagement.line_items and len(engagement.line_items) > 0:
            weeks = self._generate_weeks_from_line_items(engagement.line_items)
        else:
            # Default to 1 year from today when no line items
            from datetime import datetime
            today = datetime.now().date()
            start = today
            end = date(today.year + 1, today.month, today.day)
            weeks = self._generate_weeks_from_dates(start, end)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resource Plan Data"
        
        # Create metadata sheet
        metadata_ws = wb.create_sheet("Metadata")
        metadata_ws.sheet_state = "hidden"
        
        # Write metadata
        self._write_metadata(metadata_ws, engagement, opportunity, weeks, all_delivery_centers, filtered_roles, all_employees)
        
        # Write headers (simplified - no Opportunity date constraints)
        self._write_headers(ws, engagement.phases, weeks, opportunity.default_currency or "USD")
        
        # Write data rows
        min_rows = 20
        actual_num_rows = len(engagement.line_items)
        num_rows_to_write = max(actual_num_rows, min_rows)
        
        self._write_data_rows(ws, engagement.line_items, weeks, len(engagement.phases) if engagement.phases else 0, min_rows=min_rows)
        
        # Write totals row
        self._write_totals_row(ws, num_rows_to_write, len(weeks))
        
        # Apply data validation
        self._apply_validation(ws, all_delivery_centers, filtered_roles, all_employees, num_rows_to_write, len(weeks))
        
        # Create Excel Table
        self._create_excel_table(ws, num_rows_to_write, len(weeks))
        
        # Auto-size columns
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, min(25, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            if max_length > 0:
                ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)
            else:
                ws.column_dimensions[col_letter].width = 10
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    async def import_engagement_from_excel(self, engagement_id: UUID, file_path: str) -> Dict:
        """Import engagement Resource Plan from Excel file."""
        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        
        # Read metadata
        if "Metadata" not in wb.sheetnames:
            raise ValueError("Invalid template: Metadata sheet not found")
        
        metadata_ws = wb["Metadata"]
        metadata = self._read_metadata(metadata_ws)
        
        # Validate template
        if str(metadata["engagement_id"]) != str(engagement_id):
            raise ValueError(f"Template engagement_id ({metadata['engagement_id']}) does not match requested engagement_id ({engagement_id})")
        
        # Get engagement and opportunity
        engagement = await self.engagement_repo.get(engagement_id)
        if not engagement:
            raise ValueError("Engagement not found")
        
        opportunity = await self.opportunity_repo.get(engagement.opportunity_id)
        if not opportunity:
            raise ValueError("Opportunity not found")
        
        if str(metadata["opportunity_delivery_center_id"]) != str(opportunity.delivery_center_id):
            raise ValueError("Opportunity Invoice Center mismatch")
        
        # Read data sheet
        if "Resource Plan Data" not in wb.sheetnames:
            raise ValueError("Invalid template: Resource Plan Data sheet not found")
        
        data_ws = wb["Resource Plan Data"]
        
        # Validate week columns
        expected_weeks = metadata["week_start_dates"]
        actual_weeks = self._extract_week_columns(data_ws, len(expected_weeks))
        
        if len(actual_weeks) != len(expected_weeks):
            raise ValueError(f"Week column count mismatch: expected {len(expected_weeks)}, found {len(actual_weeks)}")
        
        # Parse data rows - use actual_weeks from sheet for exact column alignment
        line_items_data = self._parse_data_rows(
            data_ws,
            actual_weeks,
            metadata,
            opportunity.delivery_center_id,
            opportunity.default_currency or "USD"
        )
        
        # Upsert line items
        results = await self._upsert_line_items(engagement_id, line_items_data, actual_weeks)
        
        return results
    
    def _generate_weeks_from_line_items(self, line_items: List[EngagementLineItem]) -> List[date]:
        """Generate weeks from earliest Start Date week to latest End Date week of any resource plan role (line item)."""
        if not line_items:
            return []
        
        min_date = min(li.start_date for li in line_items)
        max_date = max(li.end_date for li in line_items)
        return self._generate_weeks_from_dates(min_date, max_date)
    
    def _generate_weeks_from_dates(self, start_date: date, end_date: date) -> List[date]:
        """Generate list of week start dates between two dates."""
        weeks = []
        current = self._get_week_start(start_date)
        end_week_start = self._get_week_start(end_date)
        
        while current <= end_week_start:
            weeks.append(current)
            current += timedelta(days=7)
        
        return weeks
    
    def _get_week_start(self, d: date) -> date:
        """Get the Sunday (week start) for a given date."""
        days_since_sunday = (d.weekday() + 1) % 7
        return d - timedelta(days=days_since_sunday)
    
    def _week_overlaps_date_range(self, week_start: date, start_date: date, end_date: date) -> bool:
        """True if week (Sun-Sat) overlaps [start_date, end_date]."""
        week_end = week_start + timedelta(days=6)
        return week_start <= end_date and week_end >= start_date
    
    def _parse_excel_date(self, value, field_name: str) -> date:
        """Parse date from Excel cell (datetime, date, or string)."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        s = str(value).strip()
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"]:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Invalid {field_name} format: {value}")
    
    def _write_metadata(self, ws, engagement, opportunity, weeks: List[date], 
                        delivery_centers, roles, employees):
        """Write metadata to hidden sheet."""
        ws["A1"] = "engagement_id"
        ws["B1"] = str(engagement.id)
        ws["A2"] = "quote_id"
        ws["B2"] = str(engagement.quote_id)
        ws["A3"] = "opportunity_id"
        ws["B3"] = str(engagement.opportunity_id)
        ws["A4"] = "opportunity_delivery_center_id"
        ws["B4"] = str(opportunity.delivery_center_id)
        ws["A5"] = "week_start_dates"
        ws["B5"] = ",".join([w.isoformat() for w in weeks])
        
        # Write phases
        if engagement.phases:
            ws["A6"] = "phases"
            for idx, phase in enumerate(engagement.phases):
                ws[f"B{6 + idx}"] = f"{phase.name}|{phase.start_date.isoformat()}|{phase.end_date.isoformat()}|{phase.color}"
        
        # Write delivery centers, roles, employees (similar to estimate export)
        start_row = 15
        ws[f"A{start_row}"] = "delivery_centers"
        for idx, dc in enumerate(delivery_centers):
            ws[f"B{start_row + idx}"] = f"{dc.id}|{dc.name}"
        
        roles_start = start_row + len(delivery_centers) + 2
        ws[f"A{roles_start}"] = "roles"
        for idx, role in enumerate(roles):
            ws[f"B{roles_start + idx}"] = f"{role.id}|{role.role_name}"
        
        employees_start = roles_start + len(roles) + 2
        ws[f"A{employees_start}"] = "employees"
        for idx, emp in enumerate(employees):
            ws[f"B{employees_start + idx}"] = f"{emp.id}|{emp.first_name} {emp.last_name}"
    
    def _write_headers(self, ws, phases: Optional[List[EngagementPhase]], weeks: List[date], currency: str):
        """Write header rows - aligned with Estimate export (phase row 1, year row 2, column headers row 3)."""
        # Row 1: Phase headers (if phases exist) - same structure as Estimate
        if phases and len(phases) > 0:
            col = 12  # Week columns start at L
            week_phases = {}
            for phase in sorted(phases, key=lambda p: p.start_date):
                phase_start = self._get_week_start(phase.start_date)
                phase_end = self._get_week_start(phase.end_date)
                for idx, week in enumerate(weeks):
                    week_end = week + timedelta(days=6)
                    if week <= phase_end and week_end >= phase_start:
                        if idx not in week_phases:
                            week_phases[idx] = []
                        week_phases[idx].append(phase)
            for week_idx, overlapping in week_phases.items():
                cell = ws.cell(row=1, column=col + week_idx)
                if len(overlapping) > 1:
                    cell.value = " / ".join(p.name for p in overlapping)
                    color1 = overlapping[0].color.replace("#", "")
                    color2 = overlapping[1].color.replace("#", "") if len(overlapping) > 1 else color1
                    if len(color1) == 6 and len(color2) == 6:
                        cell.fill = PatternFill(start_color=color1, end_color=color2, fill_type="darkUp")
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.font = Font(bold=True)
                else:
                    phase = overlapping[0]
                    cell.value = phase.name
                    color_hex = phase.color.replace("#", "")
                    if len(color_hex) == 6:
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2: Year headers for week columns (aligned with Estimate)
        col = 12
        current_year = None
        year_start_col = None
        year_ranges = []
        for idx, week in enumerate(weeks):
            if week.year != current_year:
                if current_year is not None and year_start_col is not None:
                    year_ranges.append((year_start_col, col + idx - 1, current_year))
                current_year = week.year
                year_start_col = col + idx
        if current_year is not None and year_start_col is not None:
            year_ranges.append((year_start_col, col + len(weeks) - 1, current_year))
        for start_c, end_c, year in year_ranges:
            for c in range(start_c, end_c + 1):
                cell = ws.cell(row=2, column=c)
                cell.value = year
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 3: Column headers (including week dates) - same as Estimate
        headers = [
            "Payable Center", "Role", "Employee",
            f"Cost ({currency})", f"Rate ({currency})",
            f"Cost ({currency}) Daily", f"Rate ({currency}) Daily",
            "Start Date", "End Date", "Billable", "Billable %",
        ]
        for idx, header in enumerate(headers):
            cell = ws.cell(row=3, column=idx + 1)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Week column headers
        for idx, week in enumerate(weeks):
            cell = ws.cell(row=3, column=col + idx)
            cell.value = week.strftime("%m/%d/%Y")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.protection = Protection(locked=True)

        # Totals column headers - 7 columns to match Estimate
        totals_start_col = col + len(weeks)
        total_headers = [
            "Total Hours",
            "Total Cost",
            "Total Revenue",
            "Billable Expense Amount",
            "Margin Amount",
            "Margin % (Without Expenses)",
            "Margin % (With Expenses)",
        ]
        for idx, header in enumerate(total_headers):
            cell = ws.cell(row=3, column=totals_start_col + idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.protection = Protection(locked=True)
    
    def _write_data_rows(self, ws, line_items: List[EngagementLineItem], weeks: List[date], num_phases: int, min_rows: int = 20):
        """Write data rows - aligned with Estimate (7 totals columns with formulas)."""
        start_row = 4
        week_col_start = 12
        totals_start_col = week_col_start + len(weeks)
        num_rows_to_write = max(len(line_items), min_rows)
        D, E, K = 4, 5, 11  # Cost, Rate, Billable %

        for row_idx in range(num_rows_to_write):
            row = start_row + row_idx
            line_item = line_items[row_idx] if row_idx < len(line_items) else None

            if line_item is None:
                for c in [4, 5, 6, 7]:
                    ws.cell(row=row, column=c).number_format = '#,##0.00'
                ws.cell(row=row, column=K).number_format = '0.00%'
                first_week_col = get_column_letter(week_col_start)
                last_week_col = get_column_letter(week_col_start + len(weeks) - 1)
                ws.cell(row=row, column=totals_start_col).value = f"=SUM({first_week_col}{row}:{last_week_col}{row})"
                ws.cell(row=row, column=totals_start_col).number_format = '#,##0.00'
                ws.cell(row=row, column=totals_start_col + 1).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(D)}{row}"
                ws.cell(row=row, column=totals_start_col + 1).number_format = '#,##0.00'
                ws.cell(row=row, column=totals_start_col + 2).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(E)}{row}"
                ws.cell(row=row, column=totals_start_col + 2).number_format = '#,##0.00'
                ws.cell(row=row, column=totals_start_col + 3).value = f"={get_column_letter(totals_start_col + 2)}{row}*{get_column_letter(K)}{row}"
                ws.cell(row=row, column=totals_start_col + 3).number_format = '#,##0.00'
                ws.cell(row=row, column=totals_start_col + 4).value = f"={get_column_letter(totals_start_col + 2)}{row}-{get_column_letter(totals_start_col + 1)}{row}"
                ws.cell(row=row, column=totals_start_col + 4).number_format = '#,##0.00'
                ws.cell(row=row, column=totals_start_col + 5).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,({get_column_letter(totals_start_col + 4)}{row}/{get_column_letter(totals_start_col + 2)}{row}))"
                ws.cell(row=row, column=totals_start_col + 5).number_format = '0.00%'
                ws.cell(row=row, column=totals_start_col + 6).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,(({get_column_letter(totals_start_col + 4)}{row}-{get_column_letter(totals_start_col + 3)}{row})/{get_column_letter(totals_start_col + 2)}{row}))"
                ws.cell(row=row, column=totals_start_col + 6).number_format = '0.00%'
                continue

            if hasattr(line_item, 'payable_center') and line_item.payable_center:
                ws.cell(row=row, column=1).value = line_item.payable_center.name
            elif line_item.role_rate and line_item.role_rate.delivery_center:
                ws.cell(row=row, column=1).value = line_item.role_rate.delivery_center.name
            if line_item.role_rate and line_item.role_rate.role:
                ws.cell(row=row, column=2).value = line_item.role_rate.role.role_name
            if line_item.employee:
                ws.cell(row=row, column=3).value = f"{line_item.employee.first_name} {line_item.employee.last_name}"

            ws.cell(row=row, column=4).value = float(line_item.cost)
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=5).value = float(line_item.rate)
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).value = float(line_item.cost) * 8
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            ws.cell(row=row, column=7).value = float(line_item.rate) * 8
            ws.cell(row=row, column=7).number_format = '#,##0.00'
            ws.cell(row=row, column=8).value = line_item.start_date
            ws.cell(row=row, column=9).value = line_item.end_date
            ws.cell(row=row, column=10).value = "Yes" if line_item.billable else "No"
            ws.cell(row=row, column=11).value = float(line_item.billable_expense_percentage) / 100
            ws.cell(row=row, column=11).number_format = '0.00%'

            weekly_hours_dict = {wh.week_start_date: float(wh.hours) for wh in (line_item.weekly_hours or [])}
            for week_idx, week in enumerate(weeks):
                hours = weekly_hours_dict.get(week, 0)
                ws.cell(row=row, column=week_col_start + week_idx).value = hours

            first_week_col = get_column_letter(week_col_start)
            last_week_col = get_column_letter(week_col_start + len(weeks) - 1)
            ws.cell(row=row, column=totals_start_col).value = f"=SUM({first_week_col}{row}:{last_week_col}{row})"
            ws.cell(row=row, column=totals_start_col).number_format = '#,##0.00'
            ws.cell(row=row, column=totals_start_col + 1).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(D)}{row}"
            ws.cell(row=row, column=totals_start_col + 1).number_format = '#,##0.00'
            ws.cell(row=row, column=totals_start_col + 2).value = f"={get_column_letter(totals_start_col)}{row}*{get_column_letter(E)}{row}"
            ws.cell(row=row, column=totals_start_col + 2).number_format = '#,##0.00'
            ws.cell(row=row, column=totals_start_col + 3).value = f"={get_column_letter(totals_start_col + 2)}{row}*{get_column_letter(K)}{row}"
            ws.cell(row=row, column=totals_start_col + 3).number_format = '#,##0.00'
            ws.cell(row=row, column=totals_start_col + 4).value = f"={get_column_letter(totals_start_col + 2)}{row}-{get_column_letter(totals_start_col + 1)}{row}"
            ws.cell(row=row, column=totals_start_col + 4).number_format = '#,##0.00'
            ws.cell(row=row, column=totals_start_col + 5).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,({get_column_letter(totals_start_col + 4)}{row}/{get_column_letter(totals_start_col + 2)}{row}))"
            ws.cell(row=row, column=totals_start_col + 5).number_format = '0.00%'
            ws.cell(row=row, column=totals_start_col + 6).value = f"=IF({get_column_letter(totals_start_col + 2)}{row}=0,0,(({get_column_letter(totals_start_col + 4)}{row}-{get_column_letter(totals_start_col + 3)}{row})/{get_column_letter(totals_start_col + 2)}{row}))"
            ws.cell(row=row, column=totals_start_col + 6).number_format = '0.00%'
    
    def _write_totals_row(self, ws, num_rows: int, num_weeks: int):
        """Write totals row."""
        totals_row = 4 + num_rows
        ws.cell(row=totals_row, column=1).value = "TOTALS"
        ws.cell(row=totals_row, column=1).font = Font(bold=True)
    
    def _apply_validation(self, ws, delivery_centers, roles, employees, num_rows: int, num_weeks: int):
        """Apply data validation - aligned with Estimate."""
        start_row = 4
        totals_row = start_row + num_rows
        max_validation_row = totals_row
        week_col_start = 12

        if delivery_centers:
            dc_names = [dc.name for dc in delivery_centers]
            dv = DataValidation(type="list", formula1=f'"{",".join(dc_names)}"', allow_blank=True)
            dv.add(f"A{start_row}:A{max_validation_row}")
            ws.add_data_validation(dv)
        if roles:
            role_names = [r.role_name for r in roles]
            dv = DataValidation(type="list", formula1=f'"{",".join(role_names)}"', allow_blank=True)
            dv.add(f"B{start_row}:B{max_validation_row}")
            ws.add_data_validation(dv)
        if employees:
            emp_names = [f"{e.first_name} {e.last_name}" for e in employees]
            dv = DataValidation(type="list", formula1=f'"{",".join(emp_names)}"', allow_blank=True)
            dv.add(f"C{start_row}:C{max_validation_row}")
            ws.add_data_validation(dv)

        date_dv = DataValidation(type="date", operator="between", formula1="1900-01-01", formula2="2100-12-31", allow_blank=True)
        date_dv.add(f"H{start_row}:H{max_validation_row}")
        date_dv.add(f"I{start_row}:I{max_validation_row}")
        ws.add_data_validation(date_dv)

        number_dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        number_dv.add(f"D{start_row}:E{max_validation_row}")
        number_dv.add(f"K{start_row}:K{max_validation_row}")
        ws.add_data_validation(number_dv)

        pct_dv = DataValidation(type="decimal", operator="between", formula1="0", formula2="100", allow_blank=True)
        pct_dv.add(f"K{start_row}:K{max_validation_row}")
        ws.add_data_validation(pct_dv)

        hours_dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
        for week_idx in range(num_weeks):
            col = week_col_start + week_idx
            col_letter = get_column_letter(col)
            hours_dv.add(f"{col_letter}{start_row}:{col_letter}{max_validation_row}")
        ws.add_data_validation(hours_dv)
    
    def _create_excel_table(self, ws, num_rows: int, num_weeks: int):
        """Create Excel Table - aligned with Estimate (header row 3, totals inside table, 7 totals columns)."""
        if num_rows == 0:
            return
        header_row = 3
        start_row = 4
        end_row = start_row + num_rows - 1
        totals_row = end_row + 1
        week_col_start = 12
        totals_start_col = week_col_start + num_weeks
        last_col = totals_start_col + 6  # 7 totals columns

        for col in range(1, last_col + 1):
            cell = ws.cell(row=header_row, column=col)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = f"Column{col}"
                logger.warning(f"Empty header at row {header_row}, col {col}, set to 'Column{col}'")

        table_ref = f"A{header_row}:{get_column_letter(last_col)}{totals_row}"
        for col in range(2, last_col + 1):
            cell = ws.cell(row=totals_row, column=col)
            if cell.value == "":
                cell.value = None

        table = Table(displayName="ResourcePlanTable", ref=table_ref)
        table.headerRowCount = 1
        table.totalsRowCount = 1
        auto_filter_ref = f"A{header_row}:{get_column_letter(last_col)}{end_row}"
        table.autoFilter = AutoFilter(ref=auto_filter_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

        for week_idx in range(num_weeks):
            col = week_col_start + week_idx
            col_letter = get_column_letter(col)
            cell = ws.cell(row=totals_row, column=col)
            cell.value = f"=SUBTOTAL(109,{col_letter}{start_row}:{col_letter}{end_row})"
            cell.font = Font(bold=True)

        for col_offset in range(7):
            col = totals_start_col + col_offset
            col_letter = get_column_letter(col)
            cell = ws.cell(row=totals_row, column=col)
            if col_offset == 5:
                margin_amount_col = get_column_letter(totals_start_col + 4)
                total_revenue_col = get_column_letter(totals_start_col + 2)
                cell.value = f"=IF({total_revenue_col}{totals_row}=0,0,({margin_amount_col}{totals_row}/{total_revenue_col}{totals_row}))"
                cell.number_format = '0.00%'
            elif col_offset == 6:
                margin_amount_col = get_column_letter(totals_start_col + 4)
                billable_expense_col = get_column_letter(totals_start_col + 3)
                total_revenue_col = get_column_letter(totals_start_col + 2)
                cell.value = f"=IF({total_revenue_col}{totals_row}=0,0,(({margin_amount_col}{totals_row}-{billable_expense_col}{totals_row})/{total_revenue_col}{totals_row}))"
                cell.number_format = '0.00%'
            else:
                cell.value = f"=SUBTOTAL(109,{col_letter}{start_row}:{col_letter}{end_row})"
                if col_offset in [1, 2, 3, 4]:
                    cell.number_format = '#,##0.00'
                elif col_offset == 0:
                    cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)
    
    def _read_metadata(self, ws) -> Dict:
        """Read metadata from metadata sheet (delivery_centers, roles, employees for import)."""
        metadata = {}
        for row in range(1, 200):
            key = ws[f"A{row}"].value
            value = ws[f"B{row}"].value
            if not key:
                continue
            if key == "engagement_id":
                metadata["engagement_id"] = UUID(value)
            elif key == "opportunity_delivery_center_id":
                metadata["opportunity_delivery_center_id"] = UUID(value)
            elif key == "week_start_dates":
                metadata["week_start_dates"] = [date.fromisoformat(d) for d in value.split(",") if d]
            elif key == "phases":
                metadata["phases"] = []
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    phase_str = ws[f"B{row_idx}"].value
                    parts = phase_str.split("|")
                    if len(parts) >= 4:
                        metadata["phases"].append({
                            "name": parts[0],
                            "start_date": date.fromisoformat(parts[1]),
                            "end_date": date.fromisoformat(parts[2]),
                            "color": parts[3],
                        })
                    row_idx += 1
            elif key == "delivery_centers":
                metadata["delivery_centers"] = {}
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    dc_str = ws[f"B{row_idx}"].value
                    parts = dc_str.split("|")
                    if len(parts) >= 2:
                        metadata["delivery_centers"][parts[1]] = UUID(parts[0])
                    row_idx += 1
            elif key == "roles":
                metadata["roles"] = {}
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    role_str = ws[f"B{row_idx}"].value
                    parts = role_str.split("|")
                    if len(parts) >= 2:
                        metadata["roles"][parts[1]] = UUID(parts[0])
                    row_idx += 1
            elif key == "employees":
                metadata["employees"] = {}
                row_idx = row
                while ws[f"B{row_idx}"].value:
                    emp_str = ws[f"B{row_idx}"].value
                    parts = emp_str.split("|")
                    if len(parts) >= 2:
                        emp_name = parts[1].strip()
                        metadata["employees"][emp_name] = UUID(parts[0])
                    row_idx += 1
        if "delivery_centers" not in metadata:
            metadata["delivery_centers"] = {}
        if "roles" not in metadata:
            metadata["roles"] = {}
        if "employees" not in metadata:
            metadata["employees"] = {}
        return metadata
    
    def _extract_week_columns(self, ws, expected_count: int) -> List[date]:
        """Extract week start dates from header row 3 - exact column alignment."""
        weeks = []
        col = 12
        for idx in range(expected_count):
            cell = ws.cell(row=3, column=col + idx)
            val = cell.value
            week_date = None
            if val is not None:
                if isinstance(val, datetime):
                    week_date = val.date()
                elif isinstance(val, date):
                    week_date = val
                elif isinstance(val, (int, float)) and not isinstance(val, bool):
                    base = date(1899, 12, 30)
                    week_date = base + timedelta(days=int(val))
                else:
                    s = str(val).strip()
                    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"):
                        try:
                            week_date = datetime.strptime(s, fmt).date()
                            break
                        except ValueError:
                            continue
            if week_date:
                weeks.append(self._get_week_start(week_date))
        return weeks
    
    def _parse_data_rows(self, ws, weeks: List[date], metadata: Dict, 
                         opportunity_delivery_center_id: UUID, currency: str) -> List[Dict]:
        """Parse data rows from Excel - kill & fill: Excel row N = plan row N-4."""
        line_items = []
        start_row = 4
        row = start_row
        while True:
            payable_center = ws[f"A{row}"].value
            if payable_center == "TOTALS" or (payable_center is None and row > start_row):
                break
            if payable_center is None:
                row += 1
                continue
            try:
                line_item = self._parse_line_item_row(ws, row, weeks, metadata, opportunity_delivery_center_id, currency)
                if line_item:
                    line_items.append(line_item)
            except Exception as e:
                logger.error(f"Error parsing row {row}: {e}", exc_info=True)
                raise ValueError(f"Row {row}: {e}") from e
            row += 1
        logger.info(f"Parsed {len(line_items)} line items from Resource Plan Excel (rows {start_row} to {row - 1})")
        return line_items
    
    def _parse_line_item_row(self, ws, row: int, weeks: List[date], metadata: Dict,
                             opportunity_delivery_center_id: UUID, currency: str) -> Optional[Dict]:
        """Parse a single line item row."""
        payable_center_name = ws[f"A{row}"].value
        if not payable_center_name or str(payable_center_name).strip() == "":
            return None
        if str(payable_center_name).strip().upper() == "TOTALS":
            return None
        
        delivery_centers = metadata.get("delivery_centers", {})
        delivery_center_id = delivery_centers.get(payable_center_name)
        if not delivery_center_id:
            raise ValueError(f"Invalid Payable Center '{payable_center_name}'")
        
        role_name = ws[f"B{row}"].value
        if not role_name:
            raise ValueError("Role is required")
        roles = metadata.get("roles", {})
        role_id = roles.get(role_name)
        if not role_id:
            raise ValueError(f"Invalid Role '{role_name}'")
        
        employee_id = None
        employee_name_raw = ws[f"C{row}"].value
        if employee_name_raw:
            employee_name = str(employee_name_raw).strip()
            employees = metadata.get("employees", {})
            employee_id = employees.get(employee_name)
            if not employee_id:
                for emp_name, emp_id in employees.items():
                    if emp_name.strip().lower() == employee_name.lower():
                        employee_id = emp_id
                        break
                if not employee_id:
                    normalized_excel = " ".join(employee_name.split())
                    for emp_name, emp_id in employees.items():
                        if " ".join(emp_name.split()).lower() == normalized_excel.lower():
                            employee_id = emp_id
                            break
                if not employee_id:
                    raise ValueError(f"Employee '{employee_name}' not found in metadata. Available: {list(employees.keys())[:5]}")
        
        cost_value = ws[f"D{row}"].value
        cost = Decimal(str(cost_value)) if cost_value is not None else None
        rate_value = ws[f"E{row}"].value
        rate = Decimal(str(rate_value)) if rate_value is not None else None
        
        start_date_value = ws[f"H{row}"].value
        if not start_date_value:
            raise ValueError("Start Date is required")
        start_date = self._parse_excel_date(start_date_value, "Start Date")
        
        end_date_value = ws[f"I{row}"].value
        if not end_date_value:
            raise ValueError("End Date is required")
        end_date = self._parse_excel_date(end_date_value, "End Date")
        
        if start_date > end_date:
            raise ValueError("Start Date must be <= End Date")
        
        billable_value = ws[f"J{row}"].value
        billable = str(billable_value).strip().lower() in ["yes", "true", "1", "y"] if billable_value else True
        
        billable_pct_value = ws[f"K{row}"].value
        billable_pct = Decimal("0")
        if billable_pct_value is not None:
            pct_decimal = Decimal(str(billable_pct_value))
            billable_pct = pct_decimal * 100 if pct_decimal <= 1 else pct_decimal
            if billable_pct < 0 or billable_pct > 100:
                raise ValueError("Billable % must be between 0 and 100")
        
        # Weekly hours: read each week column explicitly via ws.cell(row, col) for reliable per-row values.
        # Direct cell access avoids any iter_rows/iterator quirks. Empty = 0.
        weekly_hours = []
        week_col_start = 12
        for idx, week in enumerate(weeks):
            col = week_col_start + idx
            hours_value = ws.cell(row=row, column=col).value
            hours = Decimal("0")
            if hours_value is not None:
                if isinstance(hours_value, (int, float)) and not isinstance(hours_value, bool):
                    hours = Decimal(str(hours_value))
                elif str(hours_value).strip() != "":
                    try:
                        hours = Decimal(str(hours_value).strip())
                    except (ValueError, TypeError):
                        pass
            if hours < 0:
                raise ValueError(f"Week {week.isoformat()}: Hours must be >= 0")
            weekly_hours.append((week, hours))
        
        return {
            "delivery_center_id": delivery_center_id,
            "role_id": role_id,
            "employee_id": employee_id,
            "cost": cost,
            "rate": rate,
            "currency": currency,
            "start_date": start_date,
            "end_date": end_date,
            "billable": billable,
            "billable_expense_percentage": billable_pct,
            "weekly_hours": weekly_hours,
        }
    
    async def _upsert_line_items(self, engagement_id: UUID, line_items_data: List[Dict], weeks: List[date]) -> Dict:
        """Kill & fill: Excel row N maps to plan row N-4. Update in place, delete unmatched (skip if has timesheet entries)."""
        from app.models.role_rate import RoleRate
        from sqlalchemy import func
        
        existing_line_items = await self.line_item_repo.list_by_engagement(engagement_id)
        existing_line_items = sorted(existing_line_items, key=lambda li: li.row_order if li.row_order is not None else 999999)
        existing_by_row_order = {li.row_order: li for li in existing_line_items if li.row_order is not None}
        
        engagement = await self.engagement_repo.get(engagement_id)
        opportunity = await self.opportunity_repo.get(engagement.opportunity_id)
        opportunity_delivery_center_id = opportunity.delivery_center_id
        currency = opportunity.default_currency or "USD"
        
        created = 0
        updated = 0
        deleted = 0
        errors = []
        matched_line_item_ids = set()
        max_order = await self.line_item_repo.get_max_row_order(engagement_id)
        next_order = max_order + 1
        
        for idx, item_data in enumerate(line_items_data):
            excel_row = idx + 4
            row_order = idx
            try:
                role_rate_result = await self.session.execute(
                    select(RoleRate).where(
                        RoleRate.role_id == item_data["role_id"],
                        RoleRate.delivery_center_id == opportunity_delivery_center_id,
                        RoleRate.default_currency == item_data["currency"],
                    ).limit(1)
                )
                opportunity_role_rate = role_rate_result.scalars().first()
                if not opportunity_role_rate:
                    raise ValueError(
                        f"RoleRate not found for Role, Invoice Center, Currency. "
                        f"Create the RoleRate association first."
                    )
                
                final_cost = item_data["cost"]
                final_rate = item_data["rate"]
                if final_cost is None or final_rate is None:
                    default_rate, default_cost = await self._get_default_rates_from_role_rate(
                        opportunity_role_rate.id, item_data["role_id"], opportunity_delivery_center_id,
                        item_data["employee_id"], item_data["currency"],
                    )
                    final_rate = final_rate if final_rate is not None else default_rate
                    final_cost = final_cost if final_cost is not None else default_cost
                
                line_item = existing_by_row_order.get(row_order)
                if line_item:
                    await self.line_item_repo.update(
                        line_item.id,
                        role_rates_id=opportunity_role_rate.id,
                        payable_center_id=item_data["delivery_center_id"],
                        employee_id=item_data["employee_id"],
                        rate=final_rate, cost=final_cost,
                        start_date=item_data["start_date"], end_date=item_data["end_date"],
                        billable=item_data["billable"],
                        billable_expense_percentage=item_data["billable_expense_percentage"],
                        row_order=row_order,
                    )
                    updated += 1
                    matched_line_item_ids.add(line_item.id)
                else:
                    line_item = await self.line_item_repo.create(
                        engagement_id=engagement_id,
                        role_rates_id=opportunity_role_rate.id,
                        payable_center_id=item_data["delivery_center_id"],
                        employee_id=item_data["employee_id"],
                        rate=final_rate, cost=final_cost, currency=item_data["currency"],
                        start_date=item_data["start_date"], end_date=item_data["end_date"],
                        row_order=row_order,
                        billable=item_data["billable"],
                        billable_expense_percentage=item_data["billable_expense_percentage"],
                    )
                    created += 1
                    matched_line_item_ids.add(line_item.id)
                
                await self.session.flush()
                
                await self.weekly_hours_repo.delete_by_line_item(line_item.id)
                start_date = item_data["start_date"]
                end_date = item_data["end_date"]
                for week, hours in item_data["weekly_hours"]:
                    if self._week_overlaps_date_range(week, start_date, end_date):
                        await self.weekly_hours_repo.create(
                            engagement_line_item_id=line_item.id,
                            week_start_date=week,
                            hours=hours,
                        )
            except Exception as e:
                errors.append(f"Row {excel_row}: {str(e)}")
                logger.error(f"Row {excel_row}: {e}", exc_info=True)
        
        if line_items_data:
            from app.models.timesheet import TimesheetEntry
            all_existing_ids = {li.id for li in existing_line_items}
            unmatched_ids = all_existing_ids - matched_line_item_ids
            for line_item_id in unmatched_ids:
                try:
                    result = await self.session.execute(
                        select(func.count(TimesheetEntry.id)).where(
                            TimesheetEntry.engagement_line_item_id == line_item_id,
                        )
                    )
                    has_timesheet_entries = (result.scalar_one() or 0) > 0
                    if has_timesheet_entries:
                        errors.append(f"Cannot delete line item (has timesheet entries); row was removed from Excel")
                        logger.warning(f"Skip delete of line_item {line_item_id}: has timesheet entries")
                        continue
                    await self.weekly_hours_repo.delete_by_line_item(line_item_id)
                    await self.line_item_repo.delete(line_item_id)
                    deleted += 1
                except Exception as e:
                    errors.append(f"Failed to delete line item: {str(e)}")
        
        await self.session.commit()
        return {"created": created, "updated": updated, "deleted": deleted, "errors": errors}
    
    async def _get_default_rates_from_role_rate(
        self, role_rates_id: Optional[UUID], role_id: UUID, delivery_center_id: UUID,
        employee_id: Optional[UUID], target_currency: str,
    ) -> Tuple[Decimal, Decimal]:
        """Get default rate and cost from RoleRate, with employee override for cost."""
        role_rate = await self.role_rate_repo.get(role_rates_id) if role_rates_id else None
        if not role_rate:
            result = await self.session.execute(
                select(RoleRate).where(
                    RoleRate.role_id == role_id,
                    RoleRate.delivery_center_id == delivery_center_id,
                    RoleRate.default_currency == target_currency,
                ).limit(1)
            )
            role_rate = result.scalars().first()
        if not role_rate:
            return Decimal("0"), Decimal("0")
        rate = Decimal(str(role_rate.external_rate))
        cost = Decimal(str(role_rate.internal_cost_rate))
        rate_currency = role_rate.default_currency
        if employee_id:
            employee = await self.employee_repo.get(employee_id)
            if employee:
                centers_match = delivery_center_id == employee.delivery_center_id if delivery_center_id and employee.delivery_center_id else False
                if centers_match:
                    cost = Decimal(str(employee.internal_cost_rate))
                else:
                    emp_cost = Decimal(str(employee.internal_bill_rate))
                    emp_currency = employee.default_currency or "USD"
                    if target_currency and emp_currency.upper() != target_currency.upper():
                        emp_cost = Decimal(str(await convert_currency(float(emp_cost), emp_currency, target_currency, self.session)))
                    cost = emp_cost
        if target_currency and rate_currency and rate_currency.upper() != target_currency.upper():
            rate = Decimal(str(await convert_currency(float(rate), rate_currency, target_currency, self.session)))
            if not employee_id:
                cost = Decimal(str(await convert_currency(float(cost), rate_currency, target_currency, self.session)))
        return rate, cost
