"""Load replicon_mapping.xlsx into a lookup table (UUIDs and/or Cortex names resolved via DB)."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.account import Account
from app.models.engagement import Engagement, EngagementLineItem, EngagementPhase
from app.models.opportunity import Opportunity

logger = logging.getLogger(__name__)

# Inclusive fallback window when no line items / phases (single-candidate pick ignores dates).
_OPEN_WINDOW_START = date(1970, 1, 1)
_OPEN_WINDOW_END = date(2099, 12, 31)


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


@dataclass(frozen=True)
class MappingRecord:
    replicon_project: str
    replicon_client: str
    opportunity_id: UUID
    engagement_id: UUID | None
    phase_id: UUID | None
    account_id: UUID
    cortex_type: str  # ENGAGEMENT | SALES | HOLIDAY


@dataclass(frozen=True)
class MappingCandidate:
    """One resolved Cortex target plus DB windows for date-based disambiguation.

    ``window_start``/``window_end`` are engagement-wide (all resource-plan line items, else
    phases). ``employee_rp_windows`` holds per-employee min/max ``EngagementLineItem``
    start/end for that engagement (assigned rows only); at import time the timekeeper's
    employee id selects that window when present, otherwise the engagement-wide span is used.
    """

    record: MappingRecord
    window_start: date
    window_end: date
    employee_rp_windows: tuple[tuple[UUID, date, date], ...] = ()


@dataclass(frozen=True)
class MappingRule:
    """All Cortex targets for one (Replicon project, Replicon client) key — one or many candidates."""

    candidates: tuple[MappingCandidate, ...]


def split_mapping_cell_values(val: object) -> list[str]:
    """Split multi-value mapping cells on newlines (``\\n``, ``\\r\\n``, ``\\r``) or ``;``.

    Excel often stores one value per line inside a single cell. Empty / None -> [].
    """
    if val is None:
        return []
    s = str(val).strip()
    if not s:
        return []
    parts = re.split(r"\s*(?:\r\n|\n|\r)\s*|\s*;\s*", s)
    return [p.strip() for p in parts if p.strip()]


def effective_mapping_candidate_window(
    c: MappingCandidate, employee_id: UUID | None
) -> tuple[date, date]:
    """RP window for this timekeeper when they have an assigned line; else engagement-wide."""
    if employee_id is not None:
        for emp_id, ws, we in c.employee_rp_windows:
            if emp_id == employee_id:
                return (ws, we)
    return (c.window_start, c.window_end)


def _employee_line_covers_entry_date(
    c: MappingCandidate, employee_id: UUID, entry_date: date
) -> bool:
    """True if this engagement has an assigned RP line for ``employee_id`` covering ``entry_date``."""
    for emp_id, ws, we in c.employee_rp_windows:
        if emp_id == employee_id and ws <= entry_date <= we:
            return True
    return False


def pick_mapping_record_for_entry_date(
    rule: MappingRule,
    entry_date: date,
    employee_id: UUID | None = None,
) -> MappingRecord | None:
    """Pick the Cortex mapping row for ``entry_date`` (multi-contract) or the sole candidate.

    For multiple candidates, ``entry_date`` is tested against the employee's resource-plan
    line min/max dates when ``employee_id`` is known; otherwise the engagement-wide window.

    When several engagements still match (e.g. identical engagement-wide spans for a one-to-many
    Replicon project/client), prefer the engagement where the timekeeper already has an RP line
    covering ``entry_date`` so hours are not attributed to a sibling engagement that used
    narrowest-window / sheet-order tie-breaks and would auto-create a second line.
    """
    cands = rule.candidates
    if len(cands) == 1:
        rec = cands[0].record
        if rec.cortex_type == "ENGAGEMENT" and not rec.engagement_id:
            return None
        return rec

    matches: list[MappingCandidate] = []
    for c in cands:
        rec = c.record
        if rec.cortex_type != "ENGAGEMENT" or not rec.engagement_id:
            continue
        ws, we = effective_mapping_candidate_window(c, employee_id)
        if ws <= entry_date <= we:
            matches.append(c)

    if len(matches) == 0:
        return None
    if len(matches) == 1:
        return matches[0].record

    if employee_id is not None:
        on_employee_line = [
            c for c in matches if _employee_line_covers_entry_date(c, employee_id, entry_date)
        ]
        if len(on_employee_line) == 1:
            return on_employee_line[0].record
        if len(on_employee_line) > 1:
            matches = on_employee_line

    def span_days(c: MappingCandidate) -> int:
        ws, we = effective_mapping_candidate_window(c, employee_id)
        return (we - ws).days

    best_span = min(span_days(c) for c in matches)
    narrowed = [c for c in matches if span_days(c) == best_span]
    narrowed.sort(key=lambda c: cands.index(c))
    chosen = narrowed[0]
    if len(narrowed) > 1:
        logger.warning(
            "Multiple mapping contracts match entry_date=%s; using narrowest window then "
            "first sheet order (project=%r client=%r)",
            entry_date,
            chosen.record.replicon_project,
            chosen.record.replicon_client,
        )
    return chosen.record


def _cell_uuid(val: object) -> UUID | None:
    if val is None:
        return None
    if isinstance(val, UUID):
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        return UUID(s)
    except ValueError:
        return None


def _cell_str(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _row_cell(row_tuple: tuple[object, ...], col: dict[str, int], key: str) -> object:
    if key not in col:
        return None
    i = col[key]
    return row_tuple[i] if i < len(row_tuple) else None


def _find_col_map(headers: list[str]) -> dict[str, int]:
    """Map canonical keys to 0-based column index."""
    idx = {_norm(h): i for i, h in enumerate(headers)}
    aliases = {
        "replicon_project": ["replicon project", "project"],
        "replicon_client": ["replicon client", "client"],
        "cortex_opportunity": ["cortex opportunity", "cortex opportuntity", "opportunity id", "opportunity"],
        "cortex_engagement": ["cortex engagement", "engagement id", "engagement"],
        "cortex_phase": ["cortex phase", "phase id", "phase"],
        "cortex_account": ["cortex account", "account id", "account"],
        "cortex_type": ["cortex type", "type", "timesheet row type"],
    }
    out: dict[str, int] = {}
    for key, names in aliases.items():
        for n in names:
            if n in idx:
                out[key] = idx[n]
                break
    return out


async def _resolve_opportunity_id(session: AsyncSession, raw: str) -> UUID | None:
    u = _cell_uuid(raw)
    if u:
        return u
    n = _cell_str(raw)
    if not n:
        return None
    r = await session.execute(select(Opportunity.id).where(Opportunity.name == n))
    ids = list(r.scalars().all())
    if len(ids) > 1:
        logger.warning("Opportunity name %r matches %d rows; skipping mapping cell", n, len(ids))
        return None
    return ids[0] if ids else None


async def _resolve_account_id(session: AsyncSession, raw: str) -> UUID | None:
    u = _cell_uuid(raw)
    if u:
        return u
    n = _cell_str(raw)
    if not n:
        return None
    r = await session.execute(select(Account.id).where(Account.company_name == n))
    ids = list(r.scalars().all())
    if len(ids) > 1:
        logger.warning("Account company_name %r matches %d rows; skipping mapping cell", n, len(ids))
        return None
    return ids[0] if ids else None


async def _resolve_engagement_id(
    session: AsyncSession, opportunity_id: UUID, raw: str | None
) -> UUID | None:
    if raw is None:
        return None
    u = _cell_uuid(raw)
    if u:
        return u
    n = _cell_str(raw)
    if not n:
        return None
    r = await session.execute(
        select(Engagement.id).where(Engagement.opportunity_id == opportunity_id, Engagement.name == n)
    )
    ids = list(r.scalars().all())
    if len(ids) > 1:
        logger.warning(
            "Engagement name %r under opportunity %s matches %d rows; skipping mapping cell",
            n,
            opportunity_id,
            len(ids),
        )
        return None
    return ids[0] if ids else None


async def _resolve_phase_id(
    session: AsyncSession, engagement_id: UUID, raw: str | None
) -> UUID | None:
    if raw is None:
        return None
    u = _cell_uuid(raw)
    if u:
        return u
    n = _cell_str(raw)
    if not n:
        return None
    r = await session.execute(
        select(EngagementPhase.id).where(
            EngagementPhase.engagement_id == engagement_id,
            EngagementPhase.name == n,
        )
    )
    ids = list(r.scalars().all())
    if len(ids) > 1:
        logger.warning(
            "Phase name %r under engagement %s matches %d rows; skipping mapping cell",
            n,
            engagement_id,
            len(ids),
        )
        return None
    return ids[0] if ids else None


async def _batch_engagement_windows(
    session: AsyncSession, engagement_ids: list[UUID]
) -> dict[UUID, tuple[date, date] | None]:
    """Min/max assignment dates from line items, else phases; None if neither exists."""
    if not engagement_ids:
        return {}
    uniq = list(dict.fromkeys(engagement_ids))
    line_rows = (
        await session.execute(
            select(
                EngagementLineItem.engagement_id,
                func.min(EngagementLineItem.start_date),
                func.max(EngagementLineItem.end_date),
            )
            .where(EngagementLineItem.engagement_id.in_(uniq))
            .group_by(EngagementLineItem.engagement_id)
        )
    ).all()
    line_map = {row[0]: (row[1], row[2]) for row in line_rows}

    phase_rows = (
        await session.execute(
            select(
                EngagementPhase.engagement_id,
                func.min(EngagementPhase.start_date),
                func.max(EngagementPhase.end_date),
            )
            .where(EngagementPhase.engagement_id.in_(uniq))
            .group_by(EngagementPhase.engagement_id)
        )
    ).all()
    phase_map = {row[0]: (row[1], row[2]) for row in phase_rows}

    out: dict[UUID, tuple[date, date] | None] = {}
    for eid in uniq:
        if eid in line_map:
            out[eid] = line_map[eid]
        elif eid in phase_map:
            out[eid] = phase_map[eid]
        else:
            out[eid] = None
    return out


async def _batch_engagement_employee_rp_windows(
    session: AsyncSession, engagement_ids: list[UUID]
) -> dict[UUID, dict[UUID, tuple[date, date]]]:
    """Per-employee resource-plan bounds: min/max ``EngagementLineItem`` dates (assigned rows only)."""
    if not engagement_ids:
        return {}
    uniq = list(dict.fromkeys(engagement_ids))
    rows = (
        await session.execute(
            select(
                EngagementLineItem.engagement_id,
                EngagementLineItem.employee_id,
                func.min(EngagementLineItem.start_date),
                func.max(EngagementLineItem.end_date),
            )
            .where(
                EngagementLineItem.engagement_id.in_(uniq),
                EngagementLineItem.employee_id.isnot(None),
            )
            .group_by(EngagementLineItem.engagement_id, EngagementLineItem.employee_id)
        )
    ).all()
    out: dict[UUID, dict[UUID, tuple[date, date]]] = {}
    for eng_id, emp_id, mn, mx in rows:
        if emp_id is None:
            continue
        bucket = out.setdefault(eng_id, {})
        bucket[emp_id] = (mn, mx)
    return out


async def load_mapping_workbook(session: AsyncSession, path: Path) -> dict[tuple[str, str], MappingRule]:
    """
    Build (Replicon Project, Replicon Client) -> :class:`MappingRule`.

    Cortex Opportunity / Engagement cells may list multiple values separated by newlines
    (``\\n``, ``\\r\\n``, ``\\r``) or semicolons (same count), paired left-to-right. For ENGAGEMENT
    with multiple pairs, each candidate has an engagement-wide window from all line items (else
    phases) plus per-employee RP windows from that employee's ``EngagementLineItem`` rows.
    :func:`pick_mapping_record_for_entry_date` uses the timekeeper's employee id when provided,
    including preferring an engagement where they already have an RP line for that entry date when
    multiple sibling engagements share the same calendar window.

    Name lookups use exact string equality (trimmed, case-sensitive); if more than one DB row
    matches a name, that segment is skipped.
    """
    if not path.is_file():
        raise FileNotFoundError(f"Mapping workbook not found: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return {}
        headers = [_cell_str(c) for c in header_row]
        col = _find_col_map(headers)
        required = ("replicon_project", "replicon_client", "cortex_opportunity", "cortex_account", "cortex_type")
        missing = [k for k in required if k not in col]
        if missing:
            raise ValueError(f"Mapping workbook missing columns: {missing}. Headers: {headers}")

        out: dict[tuple[str, str], MappingRule] = {}
        for row in rows_iter:
            if not row:
                continue
            row_tuple = tuple(row)

            rp = _cell_str(_row_cell(row_tuple, col, "replicon_project"))
            rc = _cell_str(_row_cell(row_tuple, col, "replicon_client"))
            if not rp and not rc:
                continue

            opp_cell = _row_cell(row_tuple, col, "cortex_opportunity")
            acc_raw = _row_cell(row_tuple, col, "cortex_account")
            eng_cell = _row_cell(row_tuple, col, "cortex_engagement") if "cortex_engagement" in col else None
            ph_cell = _row_cell(row_tuple, col, "cortex_phase") if "cortex_phase" in col else None

            opp_segs = split_mapping_cell_values(opp_cell)
            eng_segs = split_mapping_cell_values(eng_cell)
            ph_segs = split_mapping_cell_values(ph_cell)
            # Legacy single-opportunity rows often leave Cortex Engagement blank (same as one empty segment).
            if len(opp_segs) == 1 and not eng_segs:
                eng_segs = [""]

            acc_id = await _resolve_account_id(session, acc_raw)
            if not opp_segs or not acc_id:
                continue

            ctype = _norm(_cell_str(_row_cell(row_tuple, col, "cortex_type")))
            if "sales" in ctype:
                t = "SALES"
            elif "holiday" in ctype or "pto" in ctype:
                t = "HOLIDAY"
            else:
                t = "ENGAGEMENT"

            if t != "ENGAGEMENT":
                if len(opp_segs) > 1 or len(eng_segs) > 1:
                    logger.warning(
                        "Mapping row project=%r client=%r: multiple opportunity/engagement values "
                        "are only supported for ENGAGEMENT type; skipping row",
                        rp,
                        rc,
                    )
                    continue
                opp_id = await _resolve_opportunity_id(session, opp_segs[0])
                if not opp_id:
                    continue
                eng_raw_single = eng_segs[0] if eng_segs else ""
                eng_id = await _resolve_engagement_id(session, opp_id, eng_raw_single or None)
                ph_id: UUID | None = None
                if eng_id and ph_segs:
                    ph_raw0 = ph_segs[0] if ph_segs else None
                    if ph_raw0:
                        ph_id = await _resolve_phase_id(session, eng_id, ph_raw0)
                rec = MappingRecord(
                    replicon_project=rp,
                    replicon_client=rc,
                    opportunity_id=opp_id,
                    engagement_id=eng_id,
                    phase_id=ph_id,
                    account_id=acc_id,
                    cortex_type=t,
                )
                out[(_norm(rp), _norm(rc))] = MappingRule(
                    (MappingCandidate(rec, _OPEN_WINDOW_START, _OPEN_WINDOW_END),),
                )
                continue

            # ENGAGEMENT
            if len(opp_segs) != len(eng_segs):
                logger.warning(
                    "Mapping row project=%r client=%r: opportunity segments (%d) and engagement "
                    "segments (%d) must match; skipping row",
                    rp,
                    rc,
                    len(opp_segs),
                    len(eng_segs),
                )
                continue
            n = len(opp_segs)
            if ph_segs and len(ph_segs) not in (0, 1, n):
                logger.warning(
                    "Mapping row project=%r client=%r: phase must be empty, one value, or %d values; skipping",
                    rp,
                    rc,
                    n,
                )
                continue

            pair_records: list[tuple[MappingRecord, UUID]] = []
            for i in range(n):
                oseg = opp_segs[i]
                eseg = eng_segs[i]
                if ph_segs and len(ph_segs) == n:
                    ph_seg = ph_segs[i]
                elif ph_segs and len(ph_segs) == 1:
                    ph_seg = ph_segs[0]
                else:
                    ph_seg = ""
                opp_id = await _resolve_opportunity_id(session, oseg)
                if not opp_id:
                    continue
                eng_id = await _resolve_engagement_id(session, opp_id, eseg)
                if not eng_id:
                    logger.warning(
                        "Mapping row project=%r client=%r segment %d: unresolved engagement %r",
                        rp,
                        rc,
                        i,
                        eseg,
                    )
                    continue
                ph_id: UUID | None = None
                if ph_seg:
                    ph_id = await _resolve_phase_id(session, eng_id, ph_seg)
                rec = MappingRecord(
                    replicon_project=rp,
                    replicon_client=rc,
                    opportunity_id=opp_id,
                    engagement_id=eng_id,
                    phase_id=ph_id,
                    account_id=acc_id,
                    cortex_type=t,
                )
                pair_records.append((rec, eng_id))

            if not pair_records:
                continue

            eng_ids = [eid for _, eid in pair_records]
            wins = await _batch_engagement_windows(session, eng_ids)
            wins_emp = await _batch_engagement_employee_rp_windows(session, eng_ids)
            candidates: list[MappingCandidate] = []
            for rec, eid in pair_records:
                w = wins.get(eid)
                if w is None:
                    logger.warning(
                        "Engagement %s has no line items or phases for contract window; skipping candidate",
                        eid,
                    )
                    continue
                ws, we = w
                emp_map = wins_emp.get(eid, {})
                emp_tuple = tuple(
                    (emp_id, s, e)
                    for emp_id, (s, e) in sorted(emp_map.items(), key=lambda kv: str(kv[0]))
                )
                candidates.append(
                    MappingCandidate(
                        record=rec,
                        window_start=ws,
                        window_end=we,
                        employee_rp_windows=emp_tuple,
                    )
                )

            if not candidates:
                continue

            key = (_norm(rp), _norm(rc))
            out[key] = MappingRule(tuple(candidates))
        return out
    finally:
        wb.close()
