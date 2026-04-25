"""Load Replicon time export from Excel (manual extract), e.g. time_export_04252026.xlsx."""

from __future__ import annotations

import logging
import re
import shutil
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.integrations.replicon.models import RawTimeRow

logger = logging.getLogger(__name__)


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _cell_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                continue
    return None


def _cell_hours(val: Any) -> Decimal:
    if val is None:
        return Decimal("0")
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    s = str(val).strip().replace(",", "")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def _billing_rate_billable(name: str) -> bool:
    n = (name or "").lower()
    if "non-bill" in n or "nonbill" in n or "non bill" in n:
        return False
    if "nb " in n or n.startswith("nb ") or "not billable" in n:
        return False
    return True


STATUS_HEADER = "import status"
DETAIL_HEADER = "import detail"


@dataclass(frozen=True)
class TimeExportLoadResult:
    """Rows plus per-sheet-row parse notes (Excel row numbers are 1-based)."""

    rows: list[RawTimeRow]
    parse_notes_by_row: dict[int, str]


def load_time_export_xlsx(path: Path) -> list[RawTimeRow]:
    """Parse workbook; returns rows only. Prefer :func:`load_time_export_xlsx_detailed` for row tracking."""
    return load_time_export_xlsx_detailed(path).rows


def load_time_export_xlsx_detailed(path: Path) -> TimeExportLoadResult:
    """
    Parse a Replicon-style Excel export with per-row diagnostics.

    Expected columns (case/spacing flexible): Client Name, Project Name, Entry Date,
    User Email, Hours; optional: User Name, Billing Rate Name (for billable hint).
    """
    if not path.is_file():
        raise FileNotFoundError(f"Timesheet export workbook not found: {path}")

    parse_notes: dict[int, str] = {}
    out: list[RawTimeRow] = []

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Report"] if "Report" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return TimeExportLoadResult(rows=[], parse_notes_by_row={})
        headers = [_norm_header(str(c) if c is not None else "") for c in header_row]

        def col(*names: str) -> int | None:
            for name in names:
                key = _norm_header(name)
                try:
                    return headers.index(key)
                except ValueError:
                    continue
            return None

        i_client = col("client name", "client")
        i_project = col("project name", "project")
        i_entry = col("entry date", "date")
        i_email = col("user email", "email")
        i_hours = col("hours", "duration")
        i_billing = col("billing rate name", "billing rate", "rate name")

        if i_client is None or i_project is None or i_entry is None or i_email is None or i_hours is None:
            raise ValueError(
                f"Workbook {path} missing required columns. Found headers: {header_row!r}. "
                "Need Client Name, Project Name, Entry Date, User Email, Hours."
            )

        excel_row = 1
        for row in rows_iter:
            excel_row += 1
            if not row:
                continue
            row_tuple = tuple(row)

            def cell(idx: int, rt: tuple[Any, ...] = row_tuple) -> Any:
                return rt[idx] if idx < len(rt) else None

            email = str(cell(i_email) or "").strip().lower()
            if not email or "@" not in email:
                parse_notes[excel_row] = "Skipped: missing or invalid User Email"
                continue
            entry_d = _cell_date(cell(i_entry))
            if not entry_d:
                parse_notes[excel_row] = "Skipped: missing or invalid Entry Date"
                continue
            hrs = _cell_hours(cell(i_hours))
            if hrs <= 0:
                parse_notes[excel_row] = "Skipped: zero or negative Hours"
                continue
            client = str(cell(i_client) or "").strip()
            project = str(cell(i_project) or "").strip()
            bill_name = str(cell(i_billing) or "").strip() if i_billing is not None else ""
            billable = _billing_rate_billable(bill_name)
            out.append(
                RawTimeRow(
                    login=email,
                    entry_date=entry_d,
                    hours=hrs,
                    client_name=client,
                    project_name=project,
                    billable=billable,
                    approved=True,
                    source_excel_row=excel_row,
                )
            )
        return TimeExportLoadResult(rows=out, parse_notes_by_row=parse_notes)
    finally:
        wb.close()


def write_import_status_columns(
    source_path: Path,
    row_status: dict[int, tuple[str, str]],
    output_path: Path,
) -> Path:
    """
    Copy ``source_path`` to ``output_path`` and add ``Import Status`` / ``Import Detail`` columns.

    ``row_status`` maps 1-based Excel row number to ``(status, detail)`` (e.g. ``("Failed", "…")``).
    Overwrites existing status columns if present from a prior run.
    Returns the path actually written (may differ if ``output_path`` is not writable).
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)
    wb = load_workbook(output_path, read_only=False, data_only=False)
    try:
        ws = wb["Report"] if "Report" in wb.sheetnames else wb[wb.sheetnames[0]]
        max_col = ws.max_column or 1
        header_cells = tuple(ws.cell(row=1, column=c).value for c in range(1, max_col + 1))
        headers_lower = [_norm_header(str(c) if c is not None else "") for c in header_cells]
        try:
            status_ci = headers_lower.index(_norm_header(STATUS_HEADER))
            detail_ci = headers_lower.index(_norm_header(DETAIL_HEADER))
            status_col = status_ci + 1
            detail_col = detail_ci + 1
        except ValueError:
            ncols = max_col
            while ncols > 0 and header_cells[ncols - 1] is None:
                ncols -= 1
            status_col = max(ncols, 1) + 1
            detail_col = status_col + 1
            ws.cell(row=1, column=status_col, value="Import Status")
            ws.cell(row=1, column=detail_col, value="Import Detail")

        for row_num, (status, detail) in row_status.items():
            if row_num < 2:
                continue
            ws.cell(row=row_num, column=status_col, value=status)
            ws.cell(row=row_num, column=detail_col, value=detail)

        wb.save(output_path)
    finally:
        wb.close()
    return output_path


def resolve_row_status_output_path(
    source_path: Path,
    explicit: Path | None,
) -> Path:
    """Prefer ``explicit``; else a sibling file ``<stem>_import_status<suffix>``."""
    if explicit is not None:
        return explicit
    return source_path.parent / f"{source_path.stem}_import_status{source_path.suffix}"


def write_row_status_workbook(
    source_path: Path,
    row_status: dict[int, tuple[str, str]],
    explicit_output: Path | None,
) -> Path | None:
    """
    Write annotated copy next to the source (or ``explicit_output``), then system temp on failure.

    Returns path written, or ``None`` if there is nothing to write.
    """
    if not row_status:
        return None
    primary = resolve_row_status_output_path(source_path, explicit_output)
    try:
        return write_import_status_columns(source_path, row_status, primary)
    except OSError as e:
        logger.warning(
            "Could not write import status workbook to %s (%s); retrying under %s",
            primary,
            e,
            tempfile.gettempdir(),
        )
        fallback = Path(tempfile.gettempdir()) / f"{source_path.stem}_import_status{source_path.suffix}"
        return write_import_status_columns(source_path, row_status, fallback)
