"""Unit tests for Replicon ETL helpers."""

from datetime import date
from decimal import Decimal
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.integrations.replicon.excel_timesheet_export import (
    load_time_export_xlsx,
    load_time_export_xlsx_detailed,
    write_import_status_columns,
)
from app.integrations.replicon.mapping_workbook import MappingCandidate, MappingRecord, MappingRule
from app.integrations.replicon.models import CortexMappedRow, RawTimeRow
from app.integrations.replicon.normalize import (
    aggregate_by_week_and_entry,
    build_login_to_employee_id,
    email_local_part,
    map_raw_row,
    merge_line_item_date_bounds_for_replicon_import,
    min_max_dates_per_employee_engagement,
    should_exclude_raw,
    week_end_saturday,
    week_start_sunday,
)


def test_email_local_part_any_domain():
    assert email_local_part("Jane.Doe@readyms.com") == "jane.doe"
    assert email_local_part("x@foo.bar") == "x"


def test_week_start_sunday():
    assert week_start_sunday(date(2026, 4, 25)) == date(2026, 4, 19)  # Sat -> prior Sun


def test_week_end_saturday_same_week_as_week_start_sunday():
    d = date(2026, 4, 21)  # Tue
    assert week_start_sunday(d) == date(2026, 4, 19)
    assert week_end_saturday(d) == date(2026, 4, 25)


def test_merge_line_item_bounds_import_envelope_no_approved():
    min_d = date(2026, 4, 21)
    max_d = date(2026, 4, 22)
    ns, ne = merge_line_item_date_bounds_for_replicon_import(None, None, min_d, max_d, [])
    assert ns == week_start_sunday(min_d)
    assert ne == week_end_saturday(max_d)


def test_merge_line_item_bounds_covers_approved_week_saturday():
    """Approved timesheet week must be fully inside [new_s, new_e] (Sun–Sat)."""
    min_d = max_d = date(2024, 12, 4)  # Wed in week starting 2024-12-01
    approved = [date(2024, 12, 1)]
    ns, ne = merge_line_item_date_bounds_for_replicon_import(
        date(2024, 12, 2), date(2024, 12, 4), min_d, max_d, approved
    )
    assert ns <= date(2024, 12, 1)
    assert ne >= date(2024, 12, 7)


def test_excluded_client_project():
    assert should_exclude_raw(
        RawTimeRow("u", date(2026, 1, 2), Decimal("1"), "BJC PCL", "Other", True, True)
    )
    assert should_exclude_raw(
        RawTimeRow("u", date(2026, 1, 2), Decimal("1"), "X", "Global Accounting and Finance", True, True)
    )
    assert not should_exclude_raw(
        RawTimeRow("u", date(2026, 1, 2), Decimal("1"), "Acme", "Proj", True, True)
    )


def _single_rule(rec: MappingRecord) -> MappingRule:
    return MappingRule(
        (
            MappingCandidate(
                rec,
                date(1970, 1, 1),
                date(2099, 12, 31),
            ),
        )
    )


def test_map_raw_row_and_aggregate():
    oid = uuid4()
    aid = uuid4()
    eid = uuid4()
    pid = uuid4()
    mapping = {
        ("alpha", "client a"): _single_rule(
            MappingRecord(
                replicon_project="Alpha",
                replicon_client="Client A",
                opportunity_id=oid,
                engagement_id=eid,
                phase_id=pid,
                account_id=aid,
                cortex_type="ENGAGEMENT",
            )
        )
    }
    raw = RawTimeRow("Bob", date(2026, 4, 21), Decimal("3"), "Client A", "Alpha", True, True)
    m = map_raw_row(raw, mapping)
    assert m is not None
    assert m.cortex_type == "ENGAGEMENT"

    emp = uuid4()
    login_map = {"bob": emp}
    buckets = aggregate_by_week_and_entry([m], login_map)
    assert len(buckets) == 1
    k = next(iter(buckets))
    assert k.employee_id == emp
    assert k.week_start == week_start_sunday(date(2026, 4, 21))
    assert buckets[k].hours_by_dow[2] == Decimal("3")  # Tue


def test_min_max_dates():
    oid = uuid4()
    aid = uuid4()
    eid = uuid4()
    emp = uuid4()
    rows = [
        CortexMappedRow("x", date(2026, 1, 5), Decimal("1"), True, "ENGAGEMENT", aid, oid, eid, None),
        CortexMappedRow("x", date(2026, 2, 1), Decimal("2"), True, "ENGAGEMENT", aid, oid, eid, None),
    ]
    mm = min_max_dates_per_employee_engagement(rows, {"x": emp})
    assert mm[(emp, eid)] == (date(2026, 1, 5), date(2026, 2, 1))


def test_build_login_to_employee_id_dedup_last_wins():
    a, b = uuid4(), uuid4()
    m = build_login_to_employee_id([(a, "dup@x.com"), (b, "dup@y.com")])
    assert m["dup"] == b


def test_build_login_indexes_full_email():
    emp = uuid4()
    m = build_login_to_employee_id([(emp, "Person.Name@Company.COM")])
    assert m["person.name@company.com"] == emp
    assert m["person.name"] == emp


def test_load_time_export_xlsx(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(
        [
            "Client Name",
            "Project Name",
            "Entry Date",
            "User Name",
            "Task Name",
            "Billing Rate Name",
            "Hours",
            "Timesheet Start Date",
            "Timesheet End Date",
            "User Email",
        ]
    )
    ws.append(
        [
            "Acme Corp",
            "Alpha",
            date(2026, 4, 21),
            "Doe, Jane",
            "Task",
            "Consultant",
            4,
            date(2026, 4, 19),
            date(2026, 4, 25),
            "jane.doe@example.com",
        ]
    )
    p = tmp_path / "export.xlsx"
    wb.save(p)
    rows = load_time_export_xlsx(p)
    assert len(rows) == 1
    r = rows[0]
    assert r.login == "jane.doe@example.com"
    assert r.client_name == "Acme Corp"
    assert r.project_name == "Alpha"
    assert r.hours == Decimal("4")
    assert r.entry_date == date(2026, 4, 21)
    assert r.approved is True
    assert r.source_excel_row == 2


def test_load_time_export_xlsx_detailed_parse_note_bad_email(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(
        [
            "Client Name",
            "Project Name",
            "Entry Date",
            "User Email",
            "Hours",
        ]
    )
    ws.append(["Acme", "Alpha", date(2026, 4, 21), "not-an-email", 1])
    p = tmp_path / "bad.xlsx"
    wb.save(p)
    det = load_time_export_xlsx_detailed(p)
    assert det.rows == []
    assert 2 in det.parse_notes_by_row


def test_write_import_status_columns_roundtrip(tmp_path):
    src = tmp_path / "src.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["Client Name", "Project Name", "Entry Date", "User Email", "Hours"])
    ws.append(["X", "Y", date(2026, 1, 6), "a@b.co", 1])
    wb.save(src)
    out = tmp_path / "out.xlsx"
    write_import_status_columns(src, {2: ("Failed", "unit test reason")}, out)
    wb2 = load_workbook(out, read_only=True, data_only=True)
    try:
        wsr = wb2.active
        headers = [str(c.value).lower() if c.value else "" for c in next(wsr.iter_rows(min_row=1, max_row=1))]
        assert "import status" in headers
        assert "import detail" in headers
        i_s = headers.index("import status")
        i_d = headers.index("import detail")
        row2 = list(next(wsr.iter_rows(min_row=2, max_row=2, values_only=True)))
        assert row2[i_s] == "Failed"
        assert row2[i_d] == "unit test reason"
    finally:
        wb2.close()
